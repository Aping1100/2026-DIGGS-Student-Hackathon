"""
Shallow Foundation Bearing Capacity Analysis (Isolated Footing, TWN-112 style)

Implements:
- Input parsing (foundation, groundwater, layered soils, loads)
- Service load combinations
- Eccentricity and effective dimensions
- Inclination correction
- Equivalent effective unit weights (gamma1, gamma2)
- Bearing capacity factors (phi=0 fixed; non-zero via table interpolation)
- Strength reduction coefficients
- Ultimate / allowable bearing capacity and allowable loads
- Excel export with Inputs, ServiceLoadCombinations, BearingCapacity, Formulas, Symbols
"""

from __future__ import annotations

import io
import math
from dataclasses import dataclass
from copy import deepcopy
from typing import Any, Dict, List, Optional, Tuple

GAMMA_W = 1.0  # tf/m^3

# Unit conversion (imperial <-> metric internal)
FT_TO_M = 0.3048
M_TO_FT = 1.0 / FT_TO_M
KIP_TO_TF = 0.45359237
TF_TO_KIP = 1.0 / KIP_TO_TF
KIPFT_TO_TFM = KIP_TO_TF * FT_TO_M
TFM_TO_KIPFT = 1.0 / KIPFT_TO_TFM
KSF_TO_TFM2 = 4.88242763638305
TFM2_TO_KSF = 1.0 / KSF_TO_TFM2
PCF_TO_TFM3 = 0.01601846337396014
TFM3_TO_PCF = 1.0 / PCF_TO_TFM3
TF_TO_KN = 9.80665
KN_TO_TF = 1.0 / TF_TO_KN
TFM2_TO_KPA = TF_TO_KN
KPA_TO_TFM2 = 1.0 / TFM2_TO_KPA
TFM3_TO_KNM3 = TF_TO_KN
KNM3_TO_TFM3 = 1.0 / TFM3_TO_KNM3
KNM_TO_TFM = KN_TO_TF
TFM_TO_KNM = TF_TO_KN


@dataclass
class SoilLayer:
    z_top: float
    z_bot: float
    gamma_t: float
    drainage_type: str  # "D" or "U"
    c_prime: float
    phi_prime: float
    su: float
    soil_label: str


def _safe_float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(v)
    except Exception:
        return default


def _normalize_drainage_type(raw: Any, su: float, phi_prime: float) -> str:
    t = str(raw or "").strip().upper()
    if t in {"D", "U"}:
        return t
    if su > 0 and phi_prime <= 0:
        return "U"
    return "D"


def _parse_layers(data: Dict[str, Any], unit_system: str) -> List[SoilLayer]:
    layers_raw = data.get("layers") or []
    layers: List[SoilLayer] = []

    for i, src in enumerate(layers_raw):
        if not isinstance(src, dict):
            continue

        z_top = _safe_float(src.get("z_top", src.get("depth_from", src.get("depthFrom", 0.0))), 0.0)
        z_bot = _safe_float(src.get("z_bot", src.get("depth_to", src.get("depthTo", z_top + 1.0))), z_top + 1.0)
        if z_bot <= z_top:
            thk = _safe_float(src.get("thickness", src.get("thickness_m", 0.0)), 0.0)
            z_bot = z_top + (thk if thk > 0 else 1.0)

        gamma_t = _safe_float(src.get("gamma_t", src.get("gamma", src.get("unit_weight", 18.6))), 18.6)
        c_prime = _safe_float(src.get("c_prime", src.get("c", 0.0)), 0.0)
        phi_prime = _safe_float(src.get("phi_prime", src.get("phi", 0.0)), 0.0)
        su = _safe_float(src.get("Su", src.get("su", 0.0)), 0.0)
        drainage_type = _normalize_drainage_type(src.get("drainage_type", src.get("drainageType")), su, phi_prime)
        soil_label = str(src.get("soil", src.get("soil_type", src.get("code", f"Layer-{i+1}"))) or f"Layer-{i+1}").strip()

        if unit_system == "imperial":
            z_top *= FT_TO_M
            z_bot *= FT_TO_M
            gamma_t *= PCF_TO_TFM3
            c_prime *= KSF_TO_TFM2
            su *= KSF_TO_TFM2
        else:
            gamma_t *= KNM3_TO_TFM3
            c_prime *= KPA_TO_TFM2
            su *= KPA_TO_TFM2

        layers.append(
            SoilLayer(
                z_top=z_top,
                z_bot=z_bot,
                gamma_t=gamma_t,
                drainage_type=drainage_type,
                c_prime=c_prime,
                phi_prime=phi_prime,
                su=su,
                soil_label=soil_label,
            )
        )

    if not layers:
        layers = [
            SoilLayer(0.0, 0.8, 1.92, "D", 0.0, 32.0, 0.0, "SF"),
            SoilLayer(0.8, 2.7, 1.85, "U", 0.0, 0.0, 4.0, "CL"),
            SoilLayer(2.7, 3.3, 1.95, "D", 0.0, 32.0, 0.0, "SM"),
            SoilLayer(3.3, 5.1, 1.85, "D", 0.0, 28.0, 0.0, "ML"),
            SoilLayer(5.1, 10.0, 1.95, "D", 0.0, 32.0, 0.0, "SM"),
        ]

    layers.sort(key=lambda x: (x.z_top, x.z_bot))
    return layers


def _get_layer_at_depth(layers: List[SoilLayer], depth: float) -> Optional[SoilLayer]:
    for layer in layers:
        if layer.z_top <= depth <= layer.z_bot:
            return layer
    return None


def _gamma_prime_at_depth(layers: List[SoilLayer], depth: float, dw: float) -> float:
    layer = _get_layer_at_depth(layers, depth)
    gamma_t = layer.gamma_t if layer else 1.9
    return gamma_t if depth < dw else (gamma_t - GAMMA_W)


def _sigma_v_effective(layers: List[SoilLayer], depth: float, dw: float) -> float:
    """Integrate effective unit weight from z=0 to given depth."""
    if depth <= 0:
        return 0.0

    breaks = {0.0, depth}
    if 0.0 < dw < depth:
        breaks.add(dw)
    for layer in layers:
        if 0.0 < layer.z_top < depth:
            breaks.add(layer.z_top)
        if 0.0 < layer.z_bot < depth:
            breaks.add(layer.z_bot)

    zpts = sorted(breaks)
    sigma = 0.0
    for i in range(len(zpts) - 1):
        z0 = zpts[i]
        z1 = zpts[i + 1]
        if z1 <= z0:
            continue
        zm = 0.5 * (z0 + z1)
        sigma += _gamma_prime_at_depth(layers, zm, dw) * (z1 - z0)
    return sigma


def _average_gamma_prime(layers: List[SoilLayer], z0: float, z1: float, dw: float) -> float:
    if z1 <= z0:
        return 1.9
    dsig = _sigma_v_effective(layers, z1, dw) - _sigma_v_effective(layers, z0, dw)
    return dsig / (z1 - z0)


# TWN-112 non-zero phi factors via interpolation table (phi in deg)
# phi=0 follows explicit fixed values by spec.
_BC_FACTOR_TABLE: List[Tuple[float, float, float, float]] = [
    (0.0, 5.30, 1.00, 0.00),
    (5.0, 6.49, 1.57, 0.45),
    (10.0, 8.34, 2.47, 1.22),
    (15.0, 10.98, 3.94, 2.65),
    (20.0, 14.83, 6.40, 5.39),
    (25.0, 20.72, 10.66, 10.88),
    (30.0, 30.14, 18.40, 22.40),
    (35.0, 46.12, 33.30, 48.03),
    (40.0, 75.31, 64.20, 109.41),
    (45.0, 133.87, 134.88, 271.75),
]


def _normalize_bearing_method(raw: Any) -> str:
    m = str(raw or "").strip()
    if not m:
        return "Vesic1973"
    m2 = m.replace(" ", "").replace("-", "").replace("_", "").lower()
    if "terzaghi" in m2 or m2 in {"1943", "t"}:
        return "Terzaghi1943"
    if "meyerhof" in m2 or m2 in {"1963", "m"}:
        return "Meyerhof1963"
    if "hansen" in m2 or m2 in {"1970", "h"}:
        return "Hansen1970"
    if "vesic" in m2 or "vesić" in m2 or m2 in {"1973", "v"}:
        return "Vesic1973"
    # Backward compatibility: if older clients send "TWN112", keep it as a distinct option.
    if "twn" in m2 or "112" in m2:
        return "TWN112"
    return str(raw)


def bearing_method_display(method: str) -> str:
    m = str(method or "")
    return {
        "Terzaghi1943": "Terzaghi (1943)",
        "Meyerhof1963": "Meyerhof (1963)",
        "Hansen1970": "Hansen (1970)",
        "Vesic1973": "Vesic (1973)",
        "TWN112": "TWN-112 (table/interpolation)",
    }.get(m, m)


def _bc_factors_classical(phi_deg: float, method: str) -> Tuple[float, float, float]:
    """Return (Nc, Nq, Ngamma) for classical bearing capacity theories.

    - Nq and Nc follow the classical closed-form expressions.
    - Ngamma differs by method (primary difference across the 4 theories).
    """
    phi = float(phi_deg)
    if phi <= 0.0:
        # Classical undrained (phi=0) bearing capacity factors
        return (5.14, 1.0, 0.0)

    phir = math.radians(phi)
    tanp = math.tan(phir)
    nq = math.exp(math.pi * tanp) * (math.tan(math.radians(45.0) + phir / 2.0) ** 2)
    nc = (nq - 1.0) / tanp

    m = str(method or "")
    if m == "Terzaghi1943":
        ng = (nq - 1.0) * math.tan(math.radians(1.4 * phi))
    elif m == "Meyerhof1963":
        # Common engineering form for Meyerhof Nγ
        ng = (nq - 1.0) * math.tan(math.radians(1.4 * phi))
    elif m == "Hansen1970":
        # Hansen 1970 widely-used approximation
        ng = 1.5 * (nq - 1.0) * tanp
    elif m == "Vesic1973":
        ng = 2.0 * (nq + 1.0) * tanp
    else:
        # Fallback: use Vesic as a modern default
        ng = 2.0 * (nq + 1.0) * tanp

    return (round(nc, 6), round(nq, 6), round(ng, 6))


def get_bearing_capacity_factors(phi_f: float, method: str) -> Tuple[float, float, float]:
    """Return (Nc, Nq, Ngamma) based on selected bearing capacity method."""
    m = _normalize_bearing_method(method)
    if m == "TWN112":
        # Legacy behavior: TWN-112 table interpolation (kept for backward compatibility).
        phi = float(phi_f)
        if phi <= 0.0:
            return (5.30, 1.00, 0.00)

        if phi <= _BC_FACTOR_TABLE[0][0]:
            _, nc, nq, ng = _BC_FACTOR_TABLE[0]
            return (nc, nq, ng)
        if phi >= _BC_FACTOR_TABLE[-1][0]:
            _, nc, nq, ng = _BC_FACTOR_TABLE[-1]
            return (nc, nq, ng)

        for i in range(len(_BC_FACTOR_TABLE) - 1):
            p0, nc0, nq0, ng0 = _BC_FACTOR_TABLE[i]
            p1, nc1, nq1, ng1 = _BC_FACTOR_TABLE[i + 1]
            if p0 <= phi <= p1:
                r = (phi - p0) / (p1 - p0) if p1 > p0 else 0.0
                nc = nc0 + r * (nc1 - nc0)
                nq = nq0 + r * (nq1 - nq0)
                ng = ng0 + r * (ng1 - ng0)
                return (round(nc, 4), round(nq, 4), round(ng, 4))

        return (5.30, 1.00, 0.00)

    return _bc_factors_classical(phi_f, m)


def _correction_factors(method: str, df: float, b_prime: float, l_prime: float, phi_f: float, beta_deg: float) -> Dict[str, float]:
    """Return correction factors (shape, depth, inclination).

    - Terzaghi (1943): no correction factors (all = 1.0).
    - Meyerhof (1963): apply s, d, i factors.
    - Hansen (1970): apply s, d, i factors (base/slope factors default to 1.0 in this implementation).
    - Vesic (1973): uses Hansen correction factors; Nγ differs.
    """
    m = _normalize_bearing_method(method)
    if m == "Terzaghi1943":
        return {
            "Fcs": 1.0, "Fcd": 1.0, "Fci": 1.0,
            "Fqs": 1.0, "Fqd": 1.0, "Fqi": 1.0,
            "Fgs": 1.0, "Fgd": 1.0, "Fgi": 1.0,
        }

    b_eff = max(float(b_prime), 1e-6)
    l_eff = max(float(l_prime), 1e-6)
    df_eff = max(float(df), 0.0)
    beta = max(float(beta_deg), 0.0)

    # Shape factors (rectangular)
    br = b_eff / l_eff
    fcs = 1.0 + 0.2 * br
    fqs = 1.0 + 0.1 * br
    fgs = max(0.6, 1.0 - 0.4 * br)

    # Depth factors (common engineering forms)
    fcd = 1.0 + 0.2 * (df_eff / b_eff)
    fqd = 1.0 + 0.1 * (df_eff / b_eff)
    fgd = 1.0

    # Inclination factors
    fci = (1.0 - beta / 90.0) ** 2 if beta < 90.0 else 0.0
    fqi = (1.0 - beta / 90.0) ** 2 if beta < 90.0 else 0.0
    if phi_f <= 0.0:
        fgi = 0.0
    else:
        fgi = (1.0 - beta / float(phi_f)) ** 2 if beta < float(phi_f) else 0.0

    return {
        "Fcs": float(fcs), "Fcd": float(fcd), "Fci": float(fci),
        "Fqs": float(fqs), "Fqd": float(fqd), "Fqi": float(fqi),
        "Fgs": float(fgs), "Fgd": float(fgd), "Fgi": max(0.0, float(fgi)),
    }


def _parse_loads(data: Dict[str, Any], unit_system: str) -> Dict[str, Dict[str, float]]:
    loads: Dict[str, Dict[str, float]] = {}
    for key in ("D", "L", "W", "E"):
        src = data.get(f"load_{key}") or {}
        loads[key] = {
            "Vx": _safe_float(src.get("Vx"), 0.0),
            "Vy": _safe_float(src.get("Vy"), 0.0),
            "Pz": _safe_float(src.get("Pz"), 0.0),
            "Mx": _safe_float(src.get("Mx"), 0.0),
            "My": _safe_float(src.get("My"), 0.0),
        }
        if unit_system == "imperial":
            loads[key]["Vx"] *= KIP_TO_TF
            loads[key]["Vy"] *= KIP_TO_TF
            loads[key]["Pz"] *= KIP_TO_TF
            loads[key]["Mx"] *= KIPFT_TO_TFM
            loads[key]["My"] *= KIPFT_TO_TFM
        else:
            loads[key]["Vx"] *= KN_TO_TF
            loads[key]["Vy"] *= KN_TO_TF
            loads[key]["Pz"] *= KN_TO_TF
            loads[key]["Mx"] *= KNM_TO_TFM
            loads[key]["My"] *= KNM_TO_TFM
    return loads


def _default_load_combinations() -> List[Dict[str, Any]]:
    return [
        {"id": "LC1", "description": "1.0D + 1.0L", "factors": {"D": 1.0, "L": 1.0, "W": 0.0, "E": 0.0}},
        {"id": "LC2", "description": "1.0D + 1.0L + 1.0W", "factors": {"D": 1.0, "L": 1.0, "W": 1.0, "E": 0.0}},
        {"id": "LC3", "description": "1.0D + 1.0L - 1.0W", "factors": {"D": 1.0, "L": 1.0, "W": -1.0, "E": 0.0}},
        {"id": "LC4", "description": "1.0D + 1.0L + 1.0E", "factors": {"D": 1.0, "L": 1.0, "W": 0.0, "E": 1.0}},
        {"id": "LC5", "description": "1.0D + 1.0L - 1.0E", "factors": {"D": 1.0, "L": 1.0, "W": 0.0, "E": -1.0}},
    ]


def _normalize_combinations(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    combos = data.get("load_combinations")
    if not isinstance(combos, list) or not combos:
        return _default_load_combinations()

    out: List[Dict[str, Any]] = []
    for i, c in enumerate(combos, start=1):
        if not isinstance(c, dict):
            continue
        factors = c.get("factors") or {}
        out.append(
            {
                "id": c.get("id", f"LC{i}"),
                "description": c.get("description", c.get("note", f"LC{i}")),
                "factors": {
                    "D": _safe_float(factors.get("D"), 0.0),
                    "L": _safe_float(factors.get("L"), 0.0),
                    "W": _safe_float(factors.get("W"), 0.0),
                    "E": _safe_float(factors.get("E"), 0.0),
                },
            }
        )

    return out or _default_load_combinations()


def _convert_result_to_output_units(result_metric: Dict[str, Any], unit_system: str) -> Dict[str, Any]:
    """Convert metric-internal result to requested output units."""
    out = deepcopy(result_metric)

    if unit_system == "metric":
        meta = out.get("metadata", {})
        if "gamma_w" in meta and meta["gamma_w"] is not None:
            meta["gamma_w"] = round(float(meta["gamma_w"]) * TFM3_TO_KNM3, 6)
        meta["unit_system"] = "metric"

        for L in out.get("layers", []):
            if L.get("gamma_t") is not None:
                L["gamma_t"] = round(float(L["gamma_t"]) * TFM3_TO_KNM3, 6)
            if L.get("Su") is not None:
                L["Su"] = round(float(L["Su"]) * TFM2_TO_KPA, 6)
            if L.get("c_prime") is not None:
                L["c_prime"] = round(float(L["c_prime"]) * TFM2_TO_KPA, 6)

        for name in ("loads",):
            group = out.get(name, {})
            for _, lv in group.items():
                if not isinstance(lv, dict):
                    continue
                for k in ("Vx", "Vy", "Pz"):
                    if lv.get(k) is not None:
                        lv[k] = round(float(lv[k]) * TF_TO_KN, 6)
                for k in ("Mx", "My"):
                    if lv.get(k) is not None:
                        lv[k] = round(float(lv[k]) * TFM_TO_KNM, 6)

        for sr in out.get("service_combinations", []):
            for k in ("Vsx", "Vsy", "Psz"):
                if sr.get(k) is not None:
                    sr[k] = round(float(sr[k]) * TF_TO_KN, 6)
            for k in ("Msx", "Msy"):
                if sr.get(k) is not None:
                    sr[k] = round(float(sr[k]) * TFM_TO_KNM, 6)

        for br in out.get("bearing_rows", []):
            if br.get("error"):
                continue
            for k in ("Vsx", "Vsy", "Psz"):
                if br.get(k) is not None:
                    br[k] = round(float(br[k]) * TF_TO_KN, 6)
            for k in ("Msx", "Msy"):
                if br.get(k) is not None:
                    br[k] = round(float(br[k]) * TFM_TO_KNM, 6)
            for k in ("cf", "quf", "qa1", "qa2", "qa3"):
                if br.get(k) is not None:
                    br[k] = round(float(br[k]) * TFM2_TO_KPA, 6)
            for k in ("gamma1", "gamma2"):
                if br.get(k) is not None:
                    br[k] = round(float(br[k]) * TFM3_TO_KNM3, 6)
            for k in ("Pa1", "Pa2", "Pa3"):
                if br.get(k) is not None:
                    br[k] = round(float(br[k]) * TF_TO_KN, 6)

    elif unit_system == "imperial":
        meta = out.get("metadata", {})
        for k in ("Df", "Lx", "Ly", "cx", "cy", "ecx", "ecy", "Dw", "L", "B"):
            if k in meta and meta[k] is not None:
                meta[k] = round(float(meta[k]) * M_TO_FT, 6)
        if "gamma_w" in meta and meta["gamma_w"] is not None:
            meta["gamma_w"] = round(float(meta["gamma_w"]) * TFM3_TO_PCF, 6)
        meta["unit_system"] = "imperial"

        for L in out.get("layers", []):
            for k in ("z_top", "z_bot"):
                if L.get(k) is not None:
                    L[k] = round(float(L[k]) * M_TO_FT, 6)
            if L.get("gamma_t") is not None:
                L["gamma_t"] = round(float(L["gamma_t"]) * TFM3_TO_PCF, 6)
            if L.get("Su") is not None:
                L["Su"] = round(float(L["Su"]) * TFM2_TO_KSF, 6)
            if L.get("c_prime") is not None:
                L["c_prime"] = round(float(L["c_prime"]) * TFM2_TO_KSF, 6)

        for name in ("loads",):
            group = out.get(name, {})
            for _, lv in group.items():
                if not isinstance(lv, dict):
                    continue
                for k in ("Vx", "Vy", "Pz"):
                    if lv.get(k) is not None:
                        lv[k] = round(float(lv[k]) * TF_TO_KIP, 6)
                for k in ("Mx", "My"):
                    if lv.get(k) is not None:
                        lv[k] = round(float(lv[k]) * TFM_TO_KIPFT, 6)

        for sr in out.get("service_combinations", []):
            for k in ("Vsx", "Vsy", "Psz"):
                if sr.get(k) is not None:
                    sr[k] = round(float(sr[k]) * TF_TO_KIP, 6)
            for k in ("Msx", "Msy"):
                if sr.get(k) is not None:
                    sr[k] = round(float(sr[k]) * TFM_TO_KIPFT, 6)

        for br in out.get("bearing_rows", []):
            if br.get("error"):
                continue
            for k in ("Vsx", "Vsy", "Psz"):
                if br.get(k) is not None:
                    br[k] = round(float(br[k]) * TF_TO_KIP, 6)
            for k in ("Msx", "Msy"):
                if br.get(k) is not None:
                    br[k] = round(float(br[k]) * TFM_TO_KIPFT, 6)
            for k in ("cf", "quf", "qa1", "qa2", "qa3"):
                if br.get(k) is not None:
                    br[k] = round(float(br[k]) * TFM2_TO_KSF, 6)
            for k in ("gamma1", "gamma2"):
                if br.get(k) is not None:
                    br[k] = round(float(br[k]) * TFM3_TO_PCF, 6)
            for k in ("Lx", "Ly", "ex", "ey", "Bprime", "Lprime"):
                if br.get(k) is not None:
                    br[k] = round(float(br[k]) * M_TO_FT, 6)
            for k in ("Pa1", "Pa2", "Pa3"):
                if br.get(k) is not None:
                    br[k] = round(float(br[k]) * TF_TO_KIP, 6)
    else:
        out["metadata"]["unit_system"] = "metric"

    # Keep backward alias consistent
    out["rows"] = [
        {
            "lc_id": r.get("load_case"),
            "combo_note": r.get("description"),
            "Vsx": r.get("Vsx"),
            "Vsy": r.get("Vsy"),
            "Psz": r.get("Psz"),
            "Msx": r.get("Msx"),
            "Msy": r.get("Msy"),
            "cf": r.get("cf"),
            "phi_f": r.get("phi_f"),
            "gamma1": r.get("gamma1"),
            "gamma2": r.get("gamma2"),
            "ex": r.get("ex"),
            "ey": r.get("ey"),
            "Bprime": r.get("Bprime"),
            "Lprime": r.get("Lprime"),
            "beta": r.get("beta"),
            "Fcs": r.get("Fcs"),
            "Fcd": r.get("Fcd"),
            "Fci": r.get("Fci"),
            "quf": r.get("quf"),
            "qa1": r.get("qa1"),
            "qa2": r.get("qa2"),
            "qa3": r.get("qa3"),
            "Pa1": r.get("Pa1"),
            "Pa2": r.get("Pa2"),
            "Pa3": r.get("Pa3"),
            "error": r.get("error"),
        }
        for r in out.get("bearing_rows", [])
    ]

    return out


def run_shallow_foundation_analysis(data: Dict[str, Any]) -> Dict[str, Any]:
    unit_system = str(data.get("unit_system", "metric")).strip().lower()
    if unit_system not in {"metric", "imperial"}:
        unit_system = "metric"

    bearing_method = _normalize_bearing_method(data.get("bearing_method", data.get("method", "Vesic1973")))

    # Foundation input
    df = _safe_float(data.get("Df"), 2.2)
    lx = _safe_float(data.get("Lx"), 3.2)
    ly = _safe_float(data.get("Ly"), 2.5)
    cx = _safe_float(data.get("cx"), 0.0)
    cy = _safe_float(data.get("cy"), 0.0)
    ecx = _safe_float(data.get("ecx"), 0.1)
    ecy = _safe_float(data.get("ecy"), 0.3)
    fsb1 = _safe_float(data.get("FSb1"), 3.0)
    fsb2 = _safe_float(data.get("FSb2"), 2.0)
    fsb3 = _safe_float(data.get("FSb3"), 1.1)
    dw = _safe_float(data.get("Dw"), 1.2)

    if unit_system == "imperial":
        df *= FT_TO_M
        lx *= FT_TO_M
        ly *= FT_TO_M
        cx *= FT_TO_M
        cy *= FT_TO_M
        ecx *= FT_TO_M
        ecy *= FT_TO_M
        dw *= FT_TO_M

    if lx <= 0 or ly <= 0:
        raise ValueError("Lx and Ly must be > 0")
    if df <= 0:
        raise ValueError("Df must be > 0")

    layers = _parse_layers(data, unit_system)
    loads = _parse_loads(data, unit_system)
    combos = _normalize_combinations(data)

    # Geometry definition
    L = max(lx, ly)
    B = min(lx, ly)

    # Soil parameters at foundation base
    layer_at_df = _get_layer_at_depth(layers, df)
    if layer_at_df is None:
        layer_at_df = layers[-1]

    if layer_at_df.drainage_type == "U":
        cf = layer_at_df.su
        phi_f = 0.0
    else:
        cf = layer_at_df.c_prime
        phi_f = layer_at_df.phi_prime

    # gamma2 (above base)
    sigma_v_eff_df = _sigma_v_effective(layers, df, dw)
    gamma2 = sigma_v_eff_df / df if df > 0 else 0.0

    service_rows: List[Dict[str, Any]] = []
    bearing_rows: List[Dict[str, Any]] = []

    for combo in combos:
        f = combo["factors"]
        vsx = f["D"] * loads["D"]["Vx"] + f["L"] * loads["L"]["Vx"] + f["W"] * loads["W"]["Vx"] + f["E"] * loads["E"]["Vx"]
        vsy = f["D"] * loads["D"]["Vy"] + f["L"] * loads["L"]["Vy"] + f["W"] * loads["W"]["Vy"] + f["E"] * loads["E"]["Vy"]
        psz = f["D"] * loads["D"]["Pz"] + f["L"] * loads["L"]["Pz"] + f["W"] * loads["W"]["Pz"] + f["E"] * loads["E"]["Pz"]
        msx = f["D"] * loads["D"]["Mx"] + f["L"] * loads["L"]["Mx"] + f["W"] * loads["W"]["Mx"] + f["E"] * loads["E"]["Mx"]
        msy = f["D"] * loads["D"]["My"] + f["L"] * loads["L"]["My"] + f["W"] * loads["W"]["My"] + f["E"] * loads["E"]["My"]

        service_rows.append(
            {
                "load_case": combo["id"],
                "Vsx": round(vsx, 4),
                "Vsy": round(vsy, 4),
                "Psz": round(psz, 4),
                "Msx": round(msx, 4),
                "Msy": round(msy, 4),
                "description": combo["description"],
            }
        )

        row: Dict[str, Any] = {
            "load_case": combo["id"],
            "description": combo["description"],
            "Vsx": round(vsx, 4),
            "Vsy": round(vsy, 4),
            "Psz": round(psz, 4),
            "Msx": round(msx, 4),
            "Msy": round(msy, 4),
            "cf": round(cf, 4),
            "phi_f": round(phi_f, 4),
        }

        if psz <= 0:
            row["error"] = "Psz <= 0 (invalid for eccentricity calculation)"
            bearing_rows.append(row)
            continue

        # Eccentricity
        ex = ecx + (msy / psz)
        ey = ecy - (msx / psz)

        # Effective dimensions
        lx_p = lx - 2.0 * abs(ex)
        ly_p = ly - 2.0 * abs(ey)
        if lx_p <= 0 or ly_p <= 0:
            row.update(
                {
                    "ex": round(ex, 4),
                    "ey": round(ey, 4),
                    "error": "Lx' <= 0 or Ly' <= 0 (method not applicable)",
                }
            )
            bearing_rows.append(row)
            continue

        l_prime = max(lx_p, ly_p)
        b_prime = min(lx_p, ly_p)

        # Inclination
        beta_x = math.degrees(math.atan(abs(vsx / psz)))
        beta_y = math.degrees(math.atan(abs(vsy / psz)))
        beta = max(beta_x, beta_y)

        # gamma1 below base (Df to Df+B')
        gamma1 = _average_gamma_prime(layers, df, df + b_prime, dw)

        # Bearing capacity factors
        nc, nq, ng = get_bearing_capacity_factors(phi_f, bearing_method)

        # Correction factors
        fac = _correction_factors(bearing_method, df, b_prime, l_prime, phi_f, beta)

        # Ultimate bearing capacity
        quf = (
            cf * nc * fac["Fcs"] * fac["Fcd"] * fac["Fci"]
            + gamma2 * df * nq * fac["Fqs"] * fac["Fqd"] * fac["Fqi"]
            + 0.5 * gamma1 * b_prime * ng * fac["Fgs"] * fac["Fgd"] * fac["Fgi"]
        )

        q_overburden = gamma2 * df
        qa1 = (quf - q_overburden) / fsb1 + q_overburden
        qa2 = (quf - q_overburden) / fsb2 + q_overburden
        qa3 = (quf - q_overburden) / fsb3 + q_overburden

        area_eff = b_prime * l_prime
        pa1 = qa1 * area_eff
        pa2 = qa2 * area_eff
        pa3 = qa3 * area_eff

        row.update(
            {
                "gamma1": round(gamma1, 6),
                "gamma2": round(gamma2, 6),
                "Nc": round(nc, 6),
                "Nq": round(nq, 6),
                "Ngamma": round(ng, 6),
                "Lx": round(lx, 6),
                "Ly": round(ly, 6),
                "ex": round(ex, 6),
                "ey": round(ey, 6),
                "Bprime": round(b_prime, 6),
                "Lprime": round(l_prime, 6),
                "beta_x": round(beta_x, 6),
                "beta_y": round(beta_y, 6),
                "beta": round(beta, 6),
                "Fcs": round(fac["Fcs"], 6),
                "Fcd": round(fac["Fcd"], 6),
                "Fci": round(fac["Fci"], 6),
                "Fqs": round(fac["Fqs"], 6),
                "Fqd": round(fac["Fqd"], 6),
                "Fqi": round(fac["Fqi"], 6),
                "Fgs": round(fac["Fgs"], 6),
                "Fgd": round(fac["Fgd"], 6),
                "Fgi": round(fac["Fgi"], 6),
                "quf": round(quf, 6),
                "qa1": round(qa1, 6),
                "qa2": round(qa2, 6),
                "qa3": round(qa3, 6),
                "Pa1": round(pa1, 6),
                "Pa2": round(pa2, 6),
                "Pa3": round(pa3, 6),
            }
        )
        bearing_rows.append(row)

    # Backward-compatible alias for frontend that reads data.rows
    rows_alias = [
        {
            "lc_id": r.get("load_case"),
            "combo_note": r.get("description"),
            "Vsx": r.get("Vsx"),
            "Vsy": r.get("Vsy"),
            "Psz": r.get("Psz"),
            "Msx": r.get("Msx"),
            "Msy": r.get("Msy"),
            "cf": r.get("cf"),
            "phi_f": r.get("phi_f"),
            "gamma1": r.get("gamma1"),
            "gamma2": r.get("gamma2"),
            "ex": r.get("ex"),
            "ey": r.get("ey"),
            "Bprime": r.get("Bprime"),
            "Lprime": r.get("Lprime"),
            "beta": r.get("beta"),
            "Fcs": r.get("Fcs"),
            "Fcd": r.get("Fcd"),
            "Fci": r.get("Fci"),
            "quf": r.get("quf"),
            "qa1": r.get("qa1"),
            "qa2": r.get("qa2"),
            "qa3": r.get("qa3"),
            "Pa1": r.get("Pa1"),
            "Pa2": r.get("Pa2"),
            "Pa3": r.get("Pa3"),
            "error": r.get("error"),
        }
        for r in bearing_rows
    ]

    result_metric = {
        "metadata": {
            "Df": df,
            "Lx": lx,
            "Ly": ly,
            "cx": cx,
            "cy": cy,
            "ecx": ecx,
            "ecy": ecy,
            "Dw": dw,
            "FSb1": fsb1,
            "FSb2": fsb2,
            "FSb3": fsb3,
            "L": L,
            "B": B,
            "gamma_w": GAMMA_W,
            "bearing_method": bearing_method,
            "bearing_method_display": bearing_method_display(bearing_method),
            "unit_system": "metric",
        },
        "layers": [
            {
                "z_top": Lr.z_top,
                "z_bot": Lr.z_bot,
                "gamma_t": Lr.gamma_t,
                "drainage_type": Lr.drainage_type,
                "c_prime": Lr.c_prime,
                "phi_prime": Lr.phi_prime,
                "Su": Lr.su,
                "soil": Lr.soil_label,
            }
            for Lr in layers
        ],
        "loads": loads,
        "service_combinations": service_rows,
        "bearing_rows": bearing_rows,
        "rows": rows_alias,
    }
    return _convert_result_to_output_units(result_metric, unit_system)


def _apply_header_style(ws, row_idx: int, ncol: int, max_cols: int = 25):
    from openpyxl.styles import Font, PatternFill, Border, Side

    fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # light yellow
    font = Font(name="Times New Roman", bold=True, color="000000")
    thin = Side(style="thin", color="000000")
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = fill
        cell.font = font
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    # Non-table area in same row should stay white
    white_side = Side(style="thin", color="FFFFFF")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for c in range(ncol + 1, max_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = white_fill
        cell.font = Font(name="Times New Roman", color="000000")
        cell.border = Border(top=white_side, left=white_side, right=white_side, bottom=white_side)


def _apply_body_border(ws, row_idx: int, ncol: int, max_cols: int = 25):
    from openpyxl.styles import Border, Side, PatternFill, Font

    thin = Side(style="thin", color="000000")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    body_font = Font(name="Times New Roman", color="000000")
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        cell.fill = white_fill
        cell.font = body_font
    # Non-table area in same row should stay white
    white_side = Side(style="thin", color="FFFFFF")
    for c in range(ncol + 1, max_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.border = Border(top=white_side, left=white_side, right=white_side, bottom=white_side)
        cell.fill = white_fill


def _apply_white_border(ws, row_idx: int, ncol: int = 25):
    """Non-table area: white border and white fill so grid lines do not show."""
    from openpyxl.styles import Border, Side, PatternFill

    white_side = Side(style="thin", color="FFFFFF")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.border = Border(top=white_side, left=white_side, right=white_side, bottom=white_side)
        cell.fill = white_fill


def _append_blank_row(ws, ncol: int = 25):
    """Append one full blank row (all columns empty) and keep it as non-table white area."""
    ws.append([""] * ncol)
    _apply_white_border(ws, ws.max_row, ncol)


def _rich_sub(txt: str):
    """Return TextBlock with subscript formatting for Excel Symbol cell."""
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock
    return TextBlock(InlineFont(vertAlign="subscript"), txt)


def _rich_sup(txt: str):
    """Return TextBlock with superscript formatting for Excel Symbol cell."""
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock
    return TextBlock(InlineFont(vertAlign="superscript"), txt)


def generate_shallow_foundation_excel(result: Dict[str, Any]) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.cell.rich_text import CellRichText

    wb = Workbook()

    meta = result.get("metadata", {})
    unit_system = str(meta.get("unit_system", "metric")).lower()
    is_metric = unit_system != "imperial"
    length_u = "m" if is_metric else "ft"
    pressure_u = "kPa" if is_metric else "ksf"
    weight_u = "kN/m³" if is_metric else "pcf"  # unit weight: never psf (psf is pressure)
    force_u = "kN" if is_metric else "kip"
    moment_u = "kN·m" if is_metric else "kip·ft"
    loads = result.get("loads", {})
    layers = result.get("layers", [])

    # Single sheet: Inputs + Service Load Combinations + Bearing Capacity
    # Border rule: BLACK = table (section header + data rows). WHITE = non-table (sheet title, blank rows between tables).
    ws = wb.active
    ws.title = "Data"
    ws.sheet_view.showGridLines = False

    # Non-table: sheet title row (white border)
    ws.cell(row=1, column=1, value="Shallow Foundation Bearing Capacity Analysis")
    ws.merge_cells("A1:E1")
    ws["A1"].font = Font(name="Times New Roman", size=14, bold=True, color="000000")
    ws["A1"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    _apply_white_border(ws, 1, 25)
    # Blank row before first table (white border)
    _append_blank_row(ws, 25)

    # ——— Table 1: Foundation Parameters (black border: header + body) ———
    # Description and Notes merged into one column
    ws.append(["Foundation Parameters", "Value", "Unit", "Description / Note"])
    _apply_header_style(ws, ws.max_row, 4)
    foundation_symbols = [
        CellRichText("D", _rich_sub("f")),
        CellRichText("L", _rich_sub("x")),
        CellRichText("L", _rich_sub("y")),
        CellRichText("c", _rich_sub("x")),
        CellRichText("c", _rich_sub("y")),
        CellRichText("e", _rich_sub("cx")),
        CellRichText("e", _rich_sub("cy")),
        CellRichText("FS", _rich_sub("b1")),
        CellRichText("FS", _rich_sub("b2")),
        CellRichText("FS", _rich_sub("b3")),
        CellRichText("D", _rich_sub("w")),
        CellRichText("γ", _rich_sub("w")),
    ]
    foundation_data = [
        (length_u, "Embedment depth"),
        (length_u, "Footing dimension in x-direction"),
        (length_u, "Footing dimension in y-direction"),
        (length_u, "Center x"),
        (length_u, "Center y"),
        (length_u, "Geometric eccentricity in x"),
        (length_u, "Geometric eccentricity in y"),
        ("-", "Long-term load bearing safety factor"),
        ("-", "Short-term load bearing safety factor"),
        ("-", "Ultimate limit state bearing safety factor"),
        (length_u, "Groundwater depth below ground surface"),
        (weight_u, "Water unit weight (Fixed by specification)"),
    ]
    _fkeys = ["Df", "Lx", "Ly", "cx", "cy", "ecx", "ecy", "FSb1", "FSb2", "FSb3", "Dw", "gamma_w"]
    for i, (sym, (u, desc)) in enumerate(zip(foundation_symbols, foundation_data)):
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=sym)
        ws.cell(row=row, column=2, value=meta.get(_fkeys[i]))
        ws.cell(row=row, column=3, value=u)
        ws.cell(row=row, column=4, value=desc)
        _apply_body_border(ws, row, 4)

    # Blank row before next table (white border)
    _append_blank_row(ws, 25)
    # ——— Table 2: Applied Loads (black border) ———
    ws.append(["Applied Loads", "Vx", "Vy", "Pz", "Mx", "My"])
    _apply_header_style(ws, ws.max_row, 6)
    ws.append(["", f"({force_u})", f"({force_u})", f"({force_u})", f"({moment_u})", f"({moment_u})"])
    _apply_header_style(ws, ws.max_row, 6)
    for key in ("D", "L", "W", "E"):
        Ld = loads.get(key, {})
        ws.append([key, Ld.get("Vx", 0), Ld.get("Vy", 0), Ld.get("Pz", 0), Ld.get("Mx", 0), Ld.get("My", 0)])
        _apply_body_border(ws, ws.max_row, 6)

    # Blank row before next table (white border)
    _append_blank_row(ws, 25)
    # ——— Table 3: Soil Layer Data (black border) ———
    _row_sl = ws.max_row + 1
    ws.cell(row=_row_sl, column=1, value="Soil Layer Data")
    ws.cell(row=_row_sl, column=2, value=CellRichText("z", _rich_sub("top")))
    ws.cell(row=_row_sl, column=3, value=CellRichText("z", _rich_sub("bot")))
    ws.cell(row=_row_sl, column=4, value=CellRichText("γ", _rich_sub("t")))
    ws.cell(row=_row_sl, column=5, value="Drainage")
    ws.cell(row=_row_sl, column=6, value=CellRichText("S", _rich_sub("u")))
    ws.cell(row=_row_sl, column=7, value="c′")
    ws.cell(row=_row_sl, column=8, value="φ′")
    ws.cell(row=_row_sl, column=9, value="Soil Class")
    _apply_header_style(ws, _row_sl, 9)
    ws.append(["", f"({length_u})", f"({length_u})", f"({weight_u})", "", f"({pressure_u})", f"({pressure_u})", "(deg)", ""])
    _apply_header_style(ws, ws.max_row, 9)
    for Lr in layers:
        ws.append([
            "Layer",
            Lr.get("z_top"),
            Lr.get("z_bot"),
            Lr.get("gamma_t"),
            Lr.get("drainage_type"),
            Lr.get("Su"),
            Lr.get("c_prime"),
            Lr.get("phi_prime"),
            Lr.get("soil"),
        ])
        _apply_body_border(ws, ws.max_row, 9)

    # Blank row before next table (white border)
    _append_blank_row(ws, 25)
    # ——— Table 4: Service Load Combinations (black border) ———
    ws.append(["Load Case", "Vsx", "Vsy", "Psz", "Msx", "Msy", "Combination Description"])
    _apply_header_style(ws, ws.max_row, 7)
    ws.append(["", f"({force_u})", f"({force_u})", f"({force_u})", f"({moment_u})", f"({moment_u})", ""])
    _apply_header_style(ws, ws.max_row, 7)
    for r in result.get("service_combinations", []):
        ws.append([
            r.get("load_case"),
            r.get("Vsx"),
            r.get("Vsy"),
            r.get("Psz"),
            r.get("Msx"),
            r.get("Msy"),
            r.get("description"),
        ])
        _apply_body_border(ws, ws.max_row, 7)

    # Blank row before next table (white border)
    _append_blank_row(ws, 25)
    # ——— Table 5: Bearing Capacity upper (black border) ———
    _row_upper = ws.max_row + 1
    ws.cell(row=_row_upper, column=1, value="Load Case")
    ws.cell(row=_row_upper, column=2, value=CellRichText("c", _rich_sub("f")))
    ws.cell(row=_row_upper, column=3, value=CellRichText("φ", _rich_sub("f")))
    ws.cell(row=_row_upper, column=4, value=CellRichText("γ", _rich_sub("1")))
    ws.cell(row=_row_upper, column=5, value=CellRichText("γ", _rich_sub("2")))
    ws.cell(row=_row_upper, column=6, value=CellRichText("N", _rich_sub("c")))
    ws.cell(row=_row_upper, column=7, value=CellRichText("N", _rich_sub("q")))
    ws.cell(row=_row_upper, column=8, value=CellRichText("N", _rich_sub("γ")))
    ws.cell(row=_row_upper, column=9, value=CellRichText("L", _rich_sub("x")))
    ws.cell(row=_row_upper, column=10, value=CellRichText("L", _rich_sub("y")))
    ws.cell(row=_row_upper, column=11, value=CellRichText("e", _rich_sub("x")))
    ws.cell(row=_row_upper, column=12, value=CellRichText("e", _rich_sub("y")))
    ws.cell(row=_row_upper, column=13, value=CellRichText("B", _rich_sup("′")))
    ws.cell(row=_row_upper, column=14, value=CellRichText("L", _rich_sup("′")))
    ws.cell(row=_row_upper, column=15, value=CellRichText("β", _rich_sub("x")))
    ws.cell(row=_row_upper, column=16, value=CellRichText("β", _rich_sub("y")))
    _apply_header_style(ws, _row_upper, 16)
    upper_units = ["", pressure_u, "deg", weight_u, weight_u, "-", "-", "-", length_u, length_u, length_u, length_u, length_u, length_u, "deg", "deg"]
    ws.append(upper_units)
    _apply_header_style(ws, ws.max_row, 16)
    for r in result.get("bearing_rows", []):
        ws.append([
            r.get("load_case"),
            r.get("cf"),
            r.get("phi_f"),
            r.get("gamma1"),
            r.get("gamma2"),
            r.get("Nc"),
            r.get("Nq"),
            r.get("Ngamma"),
            r.get("Lx"),
            r.get("Ly"),
            r.get("ex"),
            r.get("ey"),
            r.get("Bprime"),
            r.get("Lprime"),
            r.get("beta_x"),
            r.get("beta_y"),
        ])
        _apply_body_border(ws, ws.max_row, 16)

    # Blank row before next table (white border)
    _append_blank_row(ws, 25)
    # ——— Table 6: Bearing Capacity lower (black border) ———
    _row_lower = ws.max_row + 1
    ws.cell(row=_row_lower, column=1, value="Load Case")
    for col, sym in enumerate([
        CellRichText("F", _rich_sub("cs")), CellRichText("F", _rich_sub("cd")), CellRichText("F", _rich_sub("ci")),
        CellRichText("F", _rich_sub("qs")), CellRichText("F", _rich_sub("qd")), CellRichText("F", _rich_sub("qi")),
        CellRichText("F", _rich_sub("γs")), CellRichText("F", _rich_sub("γd")), CellRichText("F", _rich_sub("γi")),
        CellRichText("q", _rich_sub("uf")), CellRichText("q", _rich_sub("a1")), CellRichText("q", _rich_sub("a2")), CellRichText("q", _rich_sub("a3")),
        CellRichText("P", _rich_sub("a1")), CellRichText("P", _rich_sub("a2")), CellRichText("P", _rich_sub("a3")),
    ], start=2):
        ws.cell(row=_row_lower, column=col, value=sym)
    _apply_header_style(ws, _row_lower, 17)
    lower_units = ["", "-", "-", "-", "-", "-", "-", "-", "-", "-", pressure_u, pressure_u, pressure_u, pressure_u, force_u, force_u, force_u]
    ws.append(lower_units)
    _apply_header_style(ws, ws.max_row, 17)
    for r in result.get("bearing_rows", []):
        ws.append([
            r.get("load_case"),
            r.get("Fcs"),
            r.get("Fcd"),
            r.get("Fci"),
            r.get("Fqs"),
            r.get("Fqd"),
            r.get("Fqi"),
            r.get("Fgs"),
            r.get("Fgd"),
            r.get("Fgi"),
            r.get("quf"),
            r.get("qa1"),
            r.get("qa2"),
            r.get("qa3"),
            r.get("Pa1"),
            r.get("Pa2"),
            r.get("Pa3"),
        ])
        _apply_body_border(ws, ws.max_row, 17)

    # Sheet 2 — Formulas (no grid lines: white borders; font Times New Roman)
    from openpyxl.cell.rich_text import CellRichText
    from openpyxl.styles import Alignment, Border, Side

    ws_fm = wb.create_sheet("Formulas")
    thin_white = Side(style="thin", color="FFFFFF")
    border_white = Border(top=thin_white, left=thin_white, right=thin_white, bottom=thin_white)
    font_tnr = Font(name="Times New Roman")

    def writeline(value, merge_cols=1):
        r = ws_fm.max_row + 1
        ws_fm.cell(row=r, column=1, value=value)
        if merge_cols > 1:
            ws_fm.merge_cells(start_row=r, start_column=1, end_row=r, end_column=merge_cols)
        for c in range(1, merge_cols + 1):
            cell = ws_fm.cell(row=r, column=c)
            cell.border = border_white
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            cell.font = font_tnr

    # Title
    ws_fm.merge_cells("A1:C1")
    _t = ws_fm.cell(row=1, column=1, value="◆ Analysis Conditions")
    _t.font = Font(name="Times New Roman", bold=True, size=14)
    _t.alignment = Alignment(horizontal="left", vertical="center")
    _t.border = border_white
    for c in range(1, 4):
        ws_fm.cell(row=1, column=c).border = border_white
    ws_fm.append([])

    # Intro
    writeline("Bearing capacity analysis uses parameters c_f, φ_f of the stratum at the foundation level.", merge_cols=3)
    ws_fm.append([])

    # Block: Initial parameters (one equation per row)
    writeline(CellRichText("L", _rich_sub("x"), " (m), L", _rich_sub("y"), " (m), D", _rich_sub("f"), " (m), e", _rich_sub("cx"), " (m), e", _rich_sub("cy"), " (m)"), merge_cols=3)
    writeline(CellRichText("L = Max(L", _rich_sub("x"), ", L", _rich_sub("y"), "),  B = Min(L", _rich_sub("x"), ", L", _rich_sub("y"), ")"), merge_cols=3)
    writeline(CellRichText("e", _rich_sub("x"), " = e", _rich_sub("cx"), " + Msy/Psz,  e", _rich_sub("y"), " = e", _rich_sub("cy"), " − Msx/Psz"), merge_cols=3)
    writeline(CellRichText("L", _rich_sup("′"), " = Max(L", _rich_sub("x"), " − 2|e", _rich_sub("x"), "|, L", _rich_sub("y"), " − 2|e", _rich_sub("y"), "|),  B", _rich_sup("′"), " = Min(L", _rich_sub("x"), " − 2|e", _rich_sub("x"), "|, L", _rich_sub("y"), " − 2|e", _rich_sub("y"), "|)"), merge_cols=3)
    writeline(CellRichText("β", _rich_sub("x"), " = |tan⁻¹(Vsx/Psz)|,  β", _rich_sub("y"), " = |tan⁻¹(Vsy/Psz)|,  β = Max(β", _rich_sub("x"), ", β", _rich_sub("y"), ")"), merge_cols=3)
    ws_fm.append([])

    # Conditional block: φ_f < 10° with three columns (A, B, C)
    _row_cond = ws_fm.max_row + 1
    ws_fm.cell(row=_row_cond, column=1, value=CellRichText("φ", _rich_sub("f"), " < 10° :"))
    ws_fm.cell(row=_row_cond, column=1).font = Font(name="Times New Roman", bold=True)
    ws_fm.merge_cells("A{}:C{}".format(_row_cond, _row_cond))
    for c in range(1, 4):
        ws_fm.cell(row=_row_cond, column=c).border = border_white
        ws_fm.cell(row=_row_cond, column=c).font = font_tnr
    r1 = _row_cond + 1
    ws_fm.cell(row=r1,   column=1, value=CellRichText("F", _rich_sub("cs"), " = 1 + 0.2×B", _rich_sup("′"), "/L", _rich_sup("′"), " (≤1.2)"))
    ws_fm.cell(row=r1+1, column=1, value=CellRichText("F", _rich_sub("qs"), " = 1.0"))
    ws_fm.cell(row=r1+2, column=1, value=CellRichText("F", _rich_sub("γ"), "s = 1.0"))
    ws_fm.cell(row=r1,   column=2, value=CellRichText("F", _rich_sub("cd"), " = 1 + 0.2×D", _rich_sub("f"), "/B (≤1.5)"))
    ws_fm.cell(row=r1+1, column=2, value=CellRichText("F", _rich_sub("qd"), " = 1.0"))
    ws_fm.cell(row=r1+2, column=2, value=CellRichText("F", _rich_sub("γ"), "d = 1.0"))
    ws_fm.cell(row=r1,   column=3, value=CellRichText("F", _rich_sub("ci"), " = (1 − β/90°)", _rich_sup("2")))
    ws_fm.cell(row=r1+1, column=3, value=CellRichText("F", _rich_sub("qi"), " = (1 − β/90°)", _rich_sup("2")))
    ws_fm.cell(row=r1+2, column=3, value=CellRichText("F", _rich_sub("γ"), "i = (1 − β/φ", _rich_sub("f"), ")", _rich_sup("2"), " (≥0)"))
    for row in range(r1, r1+3):
        for col in range(1, 4):
            ws_fm.cell(row=row, column=col).border = border_white
            ws_fm.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            ws_fm.cell(row=row, column=col).font = font_tnr
    ws_fm.append([])

    # Block: Ultimate and allowable bearing capacity (one equation per row)
    writeline(CellRichText("q", _rich_sub("uf"), " = c", _rich_sub("f"), "·N", _rich_sub("c"), "·F", _rich_sub("cs"), "·F", _rich_sub("cd"), "·F", _rich_sub("ci"), " + γ", _rich_sub("2"), "·D", _rich_sub("f"), "·N", _rich_sub("q"), "·F", _rich_sub("qs"), "·F", _rich_sub("qd"), "·F", _rich_sub("qi"), " + 0.5·γ", _rich_sub("1"), "·B", _rich_sup("′"), "·N", _rich_sub("γ"), "·F", _rich_sub("γs"), "·F", _rich_sub("γd"), "·F", _rich_sub("γi")), merge_cols=3)
    writeline(CellRichText("q", _rich_sub("a1"), " = (q", _rich_sub("uf"), " − γ", _rich_sub("2"), "·D", _rich_sub("f"), ")/FS", _rich_sub("b1"), " + γ", _rich_sub("2"), "·D", _rich_sub("f"), ",  FS", _rich_sub("b1"), " = 3"), merge_cols=3)
    writeline(CellRichText("q", _rich_sub("a2"), " = (q", _rich_sub("uf"), " − γ", _rich_sub("2"), "·D", _rich_sub("f"), ")/FS", _rich_sub("b2"), " + γ", _rich_sub("2"), "·D", _rich_sub("f"), ",  FS", _rich_sub("b2"), " = 2"), merge_cols=3)
    writeline(CellRichText("q", _rich_sub("a3"), " = (q", _rich_sub("uf"), " − γ", _rich_sub("2"), "·D", _rich_sub("f"), ")/FS", _rich_sub("b3"), " + γ", _rich_sub("2"), "·D", _rich_sub("f"), ",  FS", _rich_sub("b3"), " = 1.1"), merge_cols=3)
    writeline(CellRichText("A", _rich_sup("′"), " = B", _rich_sup("′"), "×L", _rich_sup("′")), merge_cols=3)
    writeline(CellRichText("P", _rich_sub("a1"), " = q", _rich_sub("a1"), "×A", _rich_sup("′"), ",  P", _rich_sub("a2"), " = q", _rich_sub("a2"), "×A", _rich_sup("′"), ",  P", _rich_sub("a3"), " = q", _rich_sub("a3"), "×A", _rich_sup("′")), merge_cols=3)

    # Column width so long formulas stay on one line (no wrap)
    ws_fm.column_dimensions["A"].width = 100
    ws_fm.column_dimensions["B"].width = 50
    ws_fm.column_dimensions["C"].width = 50

    # Sheet 3 — Symbols (layout and subscripts/superscripts per Liquefaction style)
    from openpyxl.cell.rich_text import CellRichText
    from openpyxl.styles import Alignment, Border, Side, PatternFill

    ws_sy = wb.create_sheet("Symbols")
    thin_side = Side(style="thin", color="000000")
    # Title row (Liquefaction-style: merged, bold, center) — white bg, black text
    ws_sy.merge_cells("A1:C1")
    _cell_title = ws_sy.cell(row=1, column=1, value="Symbol Description (Shallow Foundation)")
    _cell_title.font = Font(name="Times New Roman", bold=True, size=14, color="000000")
    _cell_title.alignment = Alignment(horizontal="center", vertical="center")
    _cell_title.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    thin_white = Side(style="thin", color="FFFFFF")
    for c in range(1, 4):
        ws_sy.cell(row=1, column=c).border = Border(top=thin_white, left=thin_white, right=thin_white, bottom=thin_white)
    ws_sy.row_dimensions[1].height = 28
    # Header row: light yellow bg, black text, black border. Description and Notes merged into one column
    _header_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    for col, label in enumerate(["Symbol", "Description", "Unit"], start=1):
        _c = ws_sy.cell(row=2, column=col, value=label)
        _c.font = Font(name="Times New Roman", bold=True, size=11, color="000000")
        _c.fill = _header_fill
        _c.alignment = Alignment(horizontal="center" if col in (1, 3) else "left", vertical="center", wrap_text=True)
        _c.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
    ws_sy.row_dimensions[2].height = 22
    # Symbol rows: desc and notes combined into Description column
    def _desc_note(desc, note):
        return f"{desc} ({note})" if note else desc

    symbol_rows = [
        ("Df", _desc_note("Embedment depth", "Foundation parameter"), length_u),
        ("Lx", _desc_note("Footing dimension in x-direction", "Foundation parameter"), length_u),
        ("Ly", _desc_note("Footing dimension in y-direction", "Foundation parameter"), length_u),
        ("ecx", _desc_note("Geometric eccentricity in x", "Input"), length_u),
        ("ecy", _desc_note("Geometric eccentricity in y", "Input"), length_u),
        ("Dw", _desc_note("Groundwater depth below ground surface", "Input"), length_u),
        (CellRichText("γ", _rich_sub("w")), _desc_note("Water unit weight", "Fixed by specification"), weight_u),
        (CellRichText("z", _rich_sub("top")), _desc_note("Layer top depth", "Soil layer data"), length_u),
        (CellRichText("z", _rich_sub("bot")), _desc_note("Layer bottom depth", "Soil layer data"), length_u),
        (CellRichText("γ", _rich_sub("t")), _desc_note("Total unit weight", "Soil layer data"), weight_u),
        ("Su", _desc_note("Undrained shear strength", "Used for U layer"), pressure_u),
        (CellRichText("c", _rich_sup("′")), _desc_note("Effective cohesion", "Used for D layer"), pressure_u),
        (CellRichText("φ", _rich_sup("′")), _desc_note("Effective friction angle", "Used for D layer"), "deg"),
        ("Vsx", _desc_note("Service horizontal load in x", "From load combinations"), force_u),
        ("Vsy", _desc_note("Service horizontal load in y", "From load combinations"), force_u),
        ("Psz", _desc_note("Service vertical load", "Must be > 0"), force_u),
        ("Msx", _desc_note("Service moment about x", "From load combinations"), moment_u),
        ("Msy", _desc_note("Service moment about y", "From load combinations"), moment_u),
        ("ex", _desc_note("Resultant eccentricity in x", "ecx + Msy/Psz"), length_u),
        ("ey", _desc_note("Resultant eccentricity in y", "ecy − Msx/Psz"), length_u),
        (CellRichText("L", _rich_sup("′")), _desc_note("Effective footing length", "max(Lx′, Ly′)"), length_u),
        (CellRichText("B", _rich_sup("′")), _desc_note("Effective footing width", "min(Lx′, Ly′)"), length_u),
        (CellRichText("β", _rich_sub("x")), _desc_note("Load inclination angle in x", "atan(|Vsx/Psz|)"), "deg"),
        (CellRichText("β", _rich_sub("y")), _desc_note("Load inclination angle in y", "atan(|Vsy/Psz|)"), "deg"),
        ("β", _desc_note("Controlling inclination angle", "max(βx, βy)"), "deg"),
        ("cf", _desc_note("Foundation-level strength parameter", "Su for U; c′ for D"), pressure_u),
        (CellRichText("φ", _rich_sub("f")), _desc_note("Foundation-level friction angle", "0 for U"), "deg"),
        (CellRichText("γ", _rich_sub("2")), _desc_note("Equivalent effective unit weight above base", "σ′v(Df)/Df"), weight_u),
        (CellRichText("γ", _rich_sub("1")), _desc_note("Equivalent effective unit weight below base", "over depth B′"), weight_u),
        (CellRichText("σ", _rich_sup("′"), _rich_sub("v")), _desc_note("Effective vertical stress", "Used in overburden / γ₂"), pressure_u),
        ("Nc", _desc_note("Bearing capacity factor", "TWN-112"), "-"),
        ("Nq", _desc_note("Bearing capacity factor", "TWN-112"), "-"),
        (CellRichText("N", _rich_sub("γ")), _desc_note("Bearing capacity factor", "TWN-112"), "-"),
        (CellRichText("F", _rich_sub("cs")), _desc_note("Strength reduction coefficient", "Shape"), "-"),
        (CellRichText("F", _rich_sub("cd")), _desc_note("Strength reduction coefficient", "Depth"), "-"),
        (CellRichText("F", _rich_sub("ci")), _desc_note("Strength reduction coefficient", "Inclination"), "-"),
        (CellRichText("F", _rich_sub("qs")), _desc_note("Strength reduction coefficient", "Shape"), "-"),
        (CellRichText("F", _rich_sub("qd")), _desc_note("Strength reduction coefficient", "Depth"), "-"),
        (CellRichText("F", _rich_sub("qi")), _desc_note("Strength reduction coefficient", "Inclination"), "-"),
        (CellRichText("F", _rich_sub("γs")), _desc_note("Strength reduction coefficient", "Shape"), "-"),
        (CellRichText("F", _rich_sub("γd")), _desc_note("Strength reduction coefficient", "Depth"), "-"),
        (CellRichText("F", _rich_sub("γi")), _desc_note("Strength reduction coefficient", "Inclination"), "-"),
        (CellRichText("q", _rich_sub("uf")), _desc_note("Ultimate bearing capacity", "Computed"), pressure_u),
        (CellRichText("q", _rich_sub("a1")), _desc_note("Allowable bearing capacity with FSb1", "Computed"), pressure_u),
        (CellRichText("q", _rich_sub("a2")), _desc_note("Allowable bearing capacity with FSb2", "Computed"), pressure_u),
        (CellRichText("q", _rich_sub("a3")), _desc_note("Allowable bearing capacity with FSb3", "Computed"), pressure_u),
        (CellRichText("P", _rich_sub("a1")), _desc_note("Allowable bearing load with FSb1", "qa1 × A′"), force_u),
        (CellRichText("P", _rich_sub("a2")), _desc_note("Allowable bearing load with FSb2", "qa2 × A′"), force_u),
        (CellRichText("P", _rich_sub("a3")), _desc_note("Allowable bearing load with FSb3", "qa3 × A′"), force_u),
    ]
    _symbol_font = Font(name="Times New Roman", size=11, color="000000")
    for idx, (sym_val, desc, unit) in enumerate(symbol_rows, start=3):
        ws_sy.cell(row=idx, column=1, value=sym_val)
        ws_sy.cell(row=idx, column=2, value=desc)
        ws_sy.cell(row=idx, column=3, value=unit)
        _white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for c in range(1, 4):
            _cell = ws_sy.cell(row=idx, column=c)
            _cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
            _cell.fill = _white
            _cell.font = _symbol_font
        ws_sy.cell(row=idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws_sy.cell(row=idx, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws_sy.cell(row=idx, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws_sy.row_dimensions[idx].height = 20
    ws_sy.column_dimensions["A"].width = 22
    ws_sy.column_dimensions["B"].width = 70
    ws_sy.column_dimensions["C"].width = 12

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
