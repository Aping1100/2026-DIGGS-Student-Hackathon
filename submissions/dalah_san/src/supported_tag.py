"""
Braced excavation retaining wall (clay wall) analysis module.

This module implements:
1) Lateral force balance safety factor FSs (iterative strength reduction)
2) Basal heave safety factor FSh (integral-simplification style workflow)
3) Multi-sheet Excel export including Symbols sheet

Primary internal/output units are tf, tf/m, tf/m^2, tf/m^3, tf-m/m, m, deg.
"""

from __future__ import annotations

import io
import math
from dataclasses import dataclass
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ------------------------------
# Constants
# ------------------------------
GAMMA_W = 1.0  # tf/m^3
KN_TO_TF = 1.0 / 9.80665
PCF_TO_TF_M3 = 0.016018
KSF_TO_TF_M2 = 4.88243
KPA_TO_TF_M2 = KN_TO_TF


@dataclass
class Layer:
    no: int
    code: str
    drainage: str  # D / U
    top: float
    bot: float
    gamma_t: float  # tf/m^3
    c_prime: float  # tf/m^2
    phi_prime: float  # deg
    su_ratio_or_su: float  # ratio Su/sigma'v or Su
    spt_n: float | None
    dw_exc: float | None
    dw_ret: float | None
    seepage_mode: str


@dataclass
class SegmentResult:
    no: int
    layer_no: int
    code: str
    d1: float
    d2: float
    drainage: str
    gamma_t: float
    c_m: float
    phi_m: float
    su1_m: float
    su2_m: float
    k_hm: float
    k_chm: float
    sigma_v1: float
    sigma_v2: float
    sigma_v1_eff: float
    sigma_v2_eff: float
    sigma_h1: float
    sigma_h2: float
    p: float
    l: float
    m: float


def _f(v: Any, default: float = 0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(v)
    except Exception:
        return default


def _clamp(v: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, v))


def _parse_ratio(v: Any, default: float = 0.0) -> float:
    if v is None:
        return default
    s = str(v).strip()
    if not s:
        return default
    if "/" in s:
        a, b = s.split("/", 1)
        try:
            av = float(a)
            bv = float(b)
            return av / bv if bv != 0 else default
        except Exception:
            return default
    try:
        return float(s)
    except Exception:
        return default


def _rankine_ka(phi_deg: float) -> float:
    phi = math.radians(phi_deg)
    return math.tan(math.pi / 4.0 - phi / 2.0) ** 2


def _rankine_kp(phi_deg: float) -> float:
    phi = math.radians(phi_deg)
    return math.tan(math.pi / 4.0 + phi / 2.0) ** 2


def _caquot_kerisel_kp_approx(phi_deg: float, delta_ratio: float = 0.5) -> float:
    """
    Approximate placeholder for Caquot-Kerisel passive coefficient.
    Keeps interface pluggable while preserving workflow and table structure.
    """
    kp_r = _rankine_kp(phi_deg)
    phi = math.radians(phi_deg)
    bump = 1.0 + 0.20 * _clamp(delta_ratio, 0.0, 1.0) * math.tan(phi)
    return kp_r * bump


def _active_k(phi_deg: float, method: str = "Rankine") -> float:
    m = (method or "Rankine").strip().lower()
    # For now, all active methods fallback to Rankine to preserve robust behavior.
    return _rankine_ka(phi_deg)


def _passive_k(phi_deg: float, method: str = "Caquot-Kerisel", delta_ratio: float = 0.5) -> float:
    m = (method or "Caquot-Kerisel").strip().lower()
    if "caquot" in m or "kerisel" in m:
        return _caquot_kerisel_kp_approx(phi_deg, delta_ratio)
    return _rankine_kp(phi_deg)


def _integrate_linear_clip_nonnegative(s1: float, s2: float, z1: float, z2: float) -> Tuple[float, float]:
    """
    Integrate linear pressure between z1,z2 with clipping at zero.
    Returns force P (tf/m) and centroid depth zc (m from GL).
    """
    h = z2 - z1
    if h <= 0:
        return 0.0, z1

    if s1 <= 0.0 and s2 <= 0.0:
        return 0.0, z1 + h / 2.0

    if s1 >= 0.0 and s2 >= 0.0:
        p = (s1 + s2) * 0.5 * h
        if (s1 + s2) <= 1e-12:
            return 0.0, z1 + h / 2.0
        y_top = h * (2.0 * s2 + s1) / (3.0 * (s1 + s2))
        return p, z1 + y_top

    # sign change
    # s(z) = s1 + (s2-s1)*t, t in [0,1]
    t0 = -s1 / (s2 - s1)
    t0 = _clamp(t0, 0.0, 1.0)
    z0 = z1 + t0 * h

    if s1 < 0 < s2:
        hp = z2 - z0
        p = 0.5 * s2 * hp
        zc = z0 + 2.0 * hp / 3.0
        return p, zc

    # s1 > 0 > s2
    hp = z0 - z1
    p = 0.5 * s1 * hp
    zc = z1 + hp / 3.0
    return p, zc


def _integrate_linear(s1: float, s2: float, z1: float, z2: float) -> Tuple[float, float]:
    """Integrate linear (can be signed) trapezoid. Returns P and centroid depth."""
    h = z2 - z1
    if h <= 0:
        return 0.0, z1
    p = (s1 + s2) * 0.5 * h
    if abs(s1 + s2) <= 1e-12:
        return p, z1 + h / 2.0
    y_top = h * (2.0 * s2 + s1) / (3.0 * (s1 + s2))
    return p, z1 + y_top


def _fmt_num(v: Any, d: int = 2) -> Any:
    if v is None:
        return "-"
    if isinstance(v, str):
        return v
    try:
        fv = float(v)
        if math.isnan(fv) or math.isinf(fv):
            return "-"
        return round(fv, d)
    except Exception:
        return "-"


class BracedClayWallEngine:
    def __init__(self, data: Dict[str, Any]):
        self.raw = data or {}

        self.standard = self.raw.get("design_code", "TWN-112 (2023)")
        self.excavation_type = self.raw.get("excavation_type", "Braced")
        self.clay_method = self.raw.get("clay_method", "total_stress")  # total_stress / effective_stress
        self.ka_method = self.raw.get("ka_method", "Rankine")
        self.kp_method = self.raw.get("kp_method", "Caquot-Kerisel")

        self.delta_a_ratio = _parse_ratio(self.raw.get("delta_a_phi", "1/3"), 1.0 / 3.0)
        self.delta_p_ratio = _parse_ratio(self.raw.get("delta_p_phi", "1/2"), 1.0 / 2.0)
        self.cw_c_ratio = _parse_ratio(self.raw.get("cw_c_ratio", "2/3"), 2.0 / 3.0)
        self.cw_su_ratio = _parse_ratio(self.raw.get("cw_su_ratio", "2/3"), 2.0 / 3.0)

        # Geometry
        self.ds = abs(_f(self.raw.get("ds", 0.0)))
        self.de = abs(_f(self.raw.get("de", abs(_f(self.raw.get("excavation_depth", 0.0))))))
        self.dl = abs(_f(self.raw.get("dl", abs(_f(self.raw.get("wall_depth", 0.0))))) )
        self.ms = _f(self.raw.get("ms", self.raw.get("external_moment", 0.0)))

        # Safety requirements
        self.fss_req = _f(self.raw.get("fssR", self.raw.get("fss_r", 1.2)), 1.2)
        self.fsh_req = _f(self.raw.get("fshR", self.raw.get("fsh_r", 1.2)), 1.2)

        # Water
        self.dw_ret_global = abs(_f(self.raw.get("water_level_active", self.raw.get("dw_ret", self.raw.get("water_level_retained", 0.0)))))
        self.dw_exc_global = abs(_f(self.raw.get("water_level_passive", self.raw.get("dw_exc", self.raw.get("water_level_excavation", 0.0)))))

        # Surcharge
        self.surcharge_mode = str(self.raw.get("surcharge_mode", "manual")).lower()
        self.surcharge_q = _f(self.raw.get("surcharge_q", self.raw.get("q", 0.0)))
        self.surcharge_sh_manual = _f(self.raw.get("surcharge_sh", self.raw.get("manual_sh", 0.0)))
        self.sub_records = self.raw.get("sub_records", self.raw.get("surcharge_records", [])) or []

        self.layers = self._build_layers(self.raw.get("layers", []))

    def _normalize_strength_units(self, gamma_t: float, c_prime: float, su_ratio_or_su: float, drainage: str) -> Tuple[float, float, float]:
        """
        Normalize to tf system.
        If input appears to be in kN/kPa (common in existing UI), convert to tf units.
        """
        g = gamma_t
        c = c_prime
        su = su_ratio_or_su

        # Unit-weight heuristic: kN/m^3 values are usually > 10; tf/m^3 usually around 1.6~2.2
        if g > 8.0:
            g = g * KN_TO_TF

        # Cohesion/su heuristic for stress (tf/m^2 normally small-single digit to tens;
        # kPa values often > 20). Convert when value likely in kPa.
        if c > 15.0:
            c = c * KPA_TO_TF_M2

        # For U layer su field may be ratio or Su.
        if drainage == "U":
            if su > 3.0:
                su = su * KPA_TO_TF_M2
        else:
            # keep as is for drained ratio field (unused)
            pass
        return g, c, su

    def _build_layers(self, layers_raw: List[Dict[str, Any]]) -> List[Layer]:
        layers: List[Layer] = []
        prev = 0.0
        for i, lr in enumerate(layers_raw, start=1):
            thickness = _f(lr.get("thickness", 0.0))
            bot = _f(lr.get("bot_depth", 0.0))
            if thickness <= 0 and bot > prev:
                thickness = bot - prev
            if thickness <= 0:
                continue

            top = prev
            bot = top + thickness
            prev = bot

            # Drainage inference
            raw_type = str(lr.get("type", lr.get("drainage", "D"))).strip().upper()
            if raw_type in ("U", "CLAY"):
                drainage = "U"
            else:
                drainage = "D"

            gamma = _f(lr.get("gamma", lr.get("gamma_t", lr.get("unit_weight", 0.0))), 0.0)
            c_prime = _f(lr.get("c", lr.get("c_prime", 0.0)), 0.0)
            phi = _f(lr.get("phi", lr.get("phi_prime", 0.0)), 0.0)
            su = _f(lr.get("su", lr.get("su_ratio", lr.get("Su", 0.0))), 0.0)
            gamma, c_prime, su = self._normalize_strength_units(gamma, c_prime, su, drainage)

            code = str(lr.get("code", lr.get("soil", f"L{i}"))).strip() or f"L{i}"
            seepage = str(lr.get("seepage", lr.get("seepage_mode", ""))).strip() or "-"
            dw_exc = lr.get("dw_exc_layer", lr.get("dw_exc", None))
            dw_ret = lr.get("dw_ret_layer", lr.get("dw_ret", None))
            spt = lr.get("spt", lr.get("spt_n", None))

            layers.append(
                Layer(
                    no=i,
                    code=code,
                    drainage=drainage,
                    top=top,
                    bot=bot,
                    gamma_t=gamma,
                    c_prime=c_prime,
                    phi_prime=phi,
                    su_ratio_or_su=su,
                    spt_n=_f(spt, 0.0) if spt not in (None, "") else None,
                    dw_exc=_f(dw_exc, 0.0) if dw_exc not in (None, "") else None,
                    dw_ret=_f(dw_ret, 0.0) if dw_ret not in (None, "") else None,
                    seepage_mode=seepage,
                )
            )
        return layers

    def _layer_dw(self, layer: Layer) -> Tuple[float, float]:
        dw_ret = layer.dw_ret if layer.dw_ret is not None else self.dw_ret_global
        dw_exc = layer.dw_exc if layer.dw_exc is not None else self.dw_exc_global
        return abs(dw_ret), abs(dw_exc)

    def _key_depths(self) -> List[float]:
        depths = {0.0, self.ds, self.de, self.dl}
        for ly in self.layers:
            depths.add(ly.top)
            depths.add(ly.bot)
            dr, de = self._layer_dw(ly)
            if dr > 0:
                depths.add(dr)
            if de > 0:
                depths.add(de)
            # split by SUB Z if provided
        for rec in self.sub_records:
            z = abs(_f(rec.get("Z", rec.get("z", 0.0))))
            if z > 0:
                depths.add(z)
        return sorted(d for d in depths if 0 <= d <= self.dl + 1e-9)

    def _split_to_segments(self) -> List[Tuple[Layer, float, float]]:
        keys = self._key_depths()
        segs: List[Tuple[Layer, float, float]] = []
        for ly in self.layers:
            for i in range(len(keys) - 1):
                a = keys[i]
                b = keys[i + 1]
                if b <= ly.top or a >= ly.bot:
                    continue
                z1 = max(a, ly.top)
                z2 = min(b, ly.bot)
                if z2 - z1 > 1e-9:
                    segs.append((ly, z1, z2))
        segs.sort(key=lambda x: (x[1], x[2]))
        return segs

    def _vertical_total_sigma(self, z: float) -> float:
        sig = 0.0
        for ly in self.layers:
            if z <= ly.top:
                break
            t = min(z, ly.bot) - ly.top
            if t > 0:
                sig += ly.gamma_t * t
        return sig

    def _u_out_in(self, layer: Layer, z: float) -> Tuple[float, float]:
        dw_ret, dw_exc = self._layer_dw(layer)
        u_out = max(0.0, z - dw_ret) * GAMMA_W
        u_in = max(0.0, z - dw_exc) * GAMMA_W

        mode = (layer.seepage_mode or "-").strip().upper()
        if mode == "H":
            dh = dw_ret - dw_exc
            # keep simple symmetric seepage increment for retained/excavation sides
            seep = -0.5 * dh * GAMMA_W
            u_out += seep
            u_in -= seep
        return u_out, u_in

    def _su_at(self, layer: Layer, sigma_v_eff: float) -> float:
        """Interpret U-layer field as ratio Su/sigma'v when <=3, else absolute Su."""
        r = layer.su_ratio_or_su
        if r <= 0:
            return 0.0
        if r <= 3.0:
            return max(0.0, r * max(0.0, sigma_v_eff))
        return r

    def _surcharge_lateral_at(self, z: float, ka_ref: float) -> float:
        mode = self.surcharge_mode
        if mode == "manual":
            return max(0.0, self.surcharge_sh_manual)
        if mode == "suc":
            return max(0.0, ka_ref * self.surcharge_q)
        if mode == "sub":
            val = 0.0
            for rec in self.sub_records:
                q = _f(rec.get("Q", rec.get("q", 0.0)), 0.0)
                a = max(1e-6, abs(_f(rec.get("A", rec.get("x1", 0.0)), 0.0)))
                b = max(1e-6, abs(_f(rec.get("B", rec.get("x2", 0.0)), 0.0)))
                z_eval = max(1e-6, z)
                term = math.atan(((b - a) * z_eval) / (a * b + z_eval * z_eval))
                term += (a * z_eval) / (a * a + z_eval * z_eval)
                term -= (b * z_eval) / (b * b + z_eval * z_eval)
                val += q / math.pi * term
            return max(0.0, val)
        return 0.0

    def _lateral_for_fs(self, fs: float) -> Dict[str, Any]:
        fs = max(1.0, fs)
        segs = self._split_to_segments()
        active_rows: List[Dict[str, Any]] = []
        passive_rows: List[Dict[str, Any]] = []
        water_rows: List[Dict[str, Any]] = []
        surcharge_rows: List[Dict[str, Any]] = []

        sum_ma = 0.0
        sum_mp = 0.0
        sum_mwa = 0.0
        sum_mwp = 0.0
        sum_mqh = 0.0

        running_no = 0
        ka_ref = _active_k(self.layers[0].phi_prime if self.layers else 30.0, self.ka_method)

        for ly, z1, z2 in segs:
            running_no += 1
            h = z2 - z1
            if h <= 0:
                continue

            sigma_v1 = self._vertical_total_sigma(z1)
            sigma_v2 = self._vertical_total_sigma(z2)

            uo1, ui1 = self._u_out_in(ly, z1)
            uo2, ui2 = self._u_out_in(ly, z2)

            sigma_v1_eff = sigma_v1 - uo1
            sigma_v2_eff = sigma_v2 - uo2

            # Strength reduction
            c_m = ly.c_prime / fs if ly.drainage == "D" else 0.0
            tan_phi_m = math.tan(math.radians(ly.phi_prime)) / fs if ly.drainage == "D" else 0.0
            phi_m = math.degrees(math.atan(tan_phi_m)) if ly.drainage == "D" else 0.0

            ka = _active_k(phi_m if ly.drainage == "D" else 0.0, self.ka_method)
            kp = _passive_k(phi_m if ly.drainage == "D" else 0.0, self.kp_method, self.delta_p_ratio)
            kachm = math.sqrt(max(ka, 0.0))
            kpchm = math.sqrt(max(kp, 0.0))

            su1 = self._su_at(ly, sigma_v1_eff) / fs if ly.drainage == "U" else 0.0
            su2 = self._su_at(ly, sigma_v2_eff) / fs if ly.drainage == "U" else 0.0

            # Active lateral pressure
            if ly.drainage == "U" and self.clay_method == "total_stress":
                sigma_a1 = sigma_v1 * ka - 2.0 * su1 * kachm
                sigma_a2 = sigma_v2 * ka - 2.0 * su2 * kachm
                c_field = 0.0
                phi_field = 0.0
            else:
                cohesion_use = c_m if ly.drainage == "D" else su1
                cohesion_use2 = c_m if ly.drainage == "D" else su2
                sigma_a1 = sigma_v1_eff * ka - 2.0 * cohesion_use * kachm
                sigma_a2 = sigma_v2_eff * ka - 2.0 * cohesion_use2 * kachm
                c_field = c_m if ly.drainage == "D" else 0.0
                phi_field = phi_m if ly.drainage == "D" else 0.0

            # Passive lateral pressure
            if ly.drainage == "U" and self.clay_method == "total_stress":
                sigma_p1 = sigma_v1 * kp + 2.0 * su1 * kpchm
                sigma_p2 = sigma_v2 * kp + 2.0 * su2 * kpchm
            else:
                cohesion_use = c_m if ly.drainage == "D" else su1
                cohesion_use2 = c_m if ly.drainage == "D" else su2
                sigma_p1 = sigma_v1_eff * kp + 2.0 * cohesion_use * kpchm
                sigma_p2 = sigma_v2_eff * kp + 2.0 * cohesion_use2 * kpchm

            # Surcharge lateral increment (driving)
            sq1 = self._surcharge_lateral_at(z1, ka_ref)
            sq2 = self._surcharge_lateral_at(z2, ka_ref)

            # Active (driving) integration with clipping
            pa, zc_a = _integrate_linear_clip_nonnegative(sigma_a1, sigma_a2, z1, z2)
            la = max(0.0, self.dl - zc_a)
            ma = pa * la

            # Passive (resisting) only below excavation depth
            pp = 0.0
            lp = 0.0
            mp = 0.0
            if z2 > self.de:
                zz1 = max(z1, self.de)
                # re-interpolate endpoint values at zz1 for passive and inside water
                t = 0.0 if z2 == z1 else (zz1 - z1) / (z2 - z1)
                sig_p1_eff = sigma_p1 + (sigma_p2 - sigma_p1) * t
                sig_p2_eff = sigma_p2
                pp, zc_p = _integrate_linear(max(0.0, sig_p1_eff), max(0.0, sig_p2_eff), zz1, z2)
                lp = max(0.0, self.dl - zc_p)
                mp = pp * lp

            # Water pressures (outside driving, inside resisting)
            pwa, zc_wa = _integrate_linear(max(0.0, uo1), max(0.0, uo2), z1, z2)
            lwa = max(0.0, self.dl - zc_wa)
            mwa = pwa * lwa

            pwp = 0.0
            lwp = 0.0
            mwp = 0.0
            if z2 > self.de:
                zz1 = max(z1, self.de)
                t = 0.0 if z2 == z1 else (zz1 - z1) / (z2 - z1)
                uin1 = max(0.0, ui1 + (ui2 - ui1) * t)
                uin2 = max(0.0, ui2)
                pwp, zc_wp = _integrate_linear(uin1, uin2, zz1, z2)
                lwp = max(0.0, self.dl - zc_wp)
                mwp = pwp * lwp

            # Surcharge integration
            pqh, zc_qh = _integrate_linear(max(0.0, sq1), max(0.0, sq2), z1, z2)
            lqh = max(0.0, self.dl - zc_qh)
            mqh = pqh * lqh

            sum_ma += ma
            sum_mp += mp
            sum_mwa += mwa
            sum_mwp += mwp
            sum_mqh += mqh

            active_rows.append({
                "No": running_no,
                "Layer": ly.no,
                "LayerID": ly.code,
                "D1": _fmt_num(z1),
                "D2": _fmt_num(z2),
                "Drainage": ly.drainage,
                "gamma_t": _fmt_num(ly.gamma_t),
                "c_m": _fmt_num(c_field),
                "phi_m": _fmt_num(phi_field),
                "Su1": _fmt_num(su1) if ly.drainage == "U" else "-",
                "Su2": _fmt_num(su2) if ly.drainage == "U" else "-",
                "Kahm": _fmt_num(ka, 4),
                "Kachm": _fmt_num(kachm, 4),
                "sigma_v1": _fmt_num(sigma_v1),
                "sigma_v2": _fmt_num(sigma_v2),
                "sigma_v1_eff": _fmt_num(sigma_v1_eff),
                "sigma_v2_eff": _fmt_num(sigma_v2_eff),
                "sigma_a1": _fmt_num(max(0.0, sigma_a1)),
                "sigma_a2": _fmt_num(max(0.0, sigma_a2)),
                "Pa": _fmt_num(pa),
                "La": _fmt_num(la),
                "Ma": _fmt_num(ma),
            })

            passive_rows.append({
                "No": running_no,
                "Layer": ly.no,
                "LayerID": ly.code,
                "D1": _fmt_num(z1),
                "D2": _fmt_num(z2),
                "Drainage": ly.drainage,
                "gamma_t": _fmt_num(ly.gamma_t),
                "c_m": _fmt_num(c_field),
                "phi_m": _fmt_num(phi_field),
                "Su1": _fmt_num(su1) if ly.drainage == "U" else "-",
                "Su2": _fmt_num(su2) if ly.drainage == "U" else "-",
                "Kphm": _fmt_num(kp, 4),
                "Kpchm": _fmt_num(kpchm, 4),
                "sigma_v1": _fmt_num(sigma_v1),
                "sigma_v2": _fmt_num(sigma_v2),
                "sigma_v1_eff": _fmt_num(sigma_v1_eff),
                "sigma_v2_eff": _fmt_num(sigma_v2_eff),
                "sigma_p1": _fmt_num(max(0.0, sigma_p1)),
                "sigma_p2": _fmt_num(max(0.0, sigma_p2)),
                "Pp": _fmt_num(pp),
                "Lp": _fmt_num(lp),
                "Mp": _fmt_num(mp),
            })

            water_rows.append({
                "No": running_no,
                "LayerID": ly.code,
                "D1": _fmt_num(z1),
                "D2": _fmt_num(z2),
                "sigma_wa1": _fmt_num(max(0.0, uo1)),
                "sigma_wa2": _fmt_num(max(0.0, uo2)),
                "sigma_wp1": _fmt_num(max(0.0, ui1)),
                "sigma_wp2": _fmt_num(max(0.0, ui2)),
                "net_1": _fmt_num(max(0.0, uo1 - ui1)),
                "net_2": _fmt_num(max(0.0, uo2 - ui2)),
                "Pwa": _fmt_num(pwa),
                "Lwa": _fmt_num(lwa),
                "Mwa": _fmt_num(mwa),
                "Pwp": _fmt_num(pwp),
                "Lwp": _fmt_num(lwp),
                "Mwp": _fmt_num(mwp),
            })

            surcharge_rows.append({
                "No": running_no,
                "Dqh1": _fmt_num(z1),
                "Dqh2": _fmt_num(z2),
                "sigma_q1": _fmt_num(max(0.0, sq1)),
                "sigma_q2": _fmt_num(max(0.0, sq2)),
                "Pqh": _fmt_num(pqh),
                "Lqh": _fmt_num(lqh),
                "Mqh": _fmt_num(mqh),
            })

        driving = sum_ma + sum_mwa + sum_mqh
        resisting = sum_mp + sum_mwp + self.ms

        # Build legacy details for compatibility with existing frontend summary view.
        details_legacy = []
        for i in range(len(active_rows)):
            a = active_rows[i]
            p = passive_rows[i]
            w = water_rows[i]
            details_legacy.append(
                {
                    "active": {
                        "D1": a["D1"],
                        "D2": a["D2"],
                        "Code": a["LayerID"],
                        "gamma_t": a["gamma_t"] if isinstance(a["gamma_t"], (int, float)) else 0.0,
                        "c_prime": a["c_m"],
                        "phi_prime": a["phi_m"],
                        "Kahm": a["Kahm"],
                        "sigma_va1_prime": a["sigma_v1_eff"],
                        "sigma_va2_prime": a["sigma_v2_eff"],
                        "sigma_a1": a["sigma_a1"],
                        "sigma_a2": a["sigma_a2"],
                        "Pa": a["Pa"],
                        "La": a["La"],
                        "Ma": a["Ma"],
                    },
                    "passive": {
                        "D1": p["D1"],
                        "D2": p["D2"],
                        "Code": p["LayerID"],
                        "gamma_t": p["gamma_t"] if isinstance(p["gamma_t"], (int, float)) else 0.0,
                        "c_prime": p["c_m"],
                        "phi_prime": p["phi_m"],
                        "Kphm": p["Kphm"],
                        "sigma_vp1_prime": p["sigma_v1_eff"],
                        "sigma_vp2_prime": p["sigma_v2_eff"],
                        "sigma_p1": p["sigma_p1"],
                        "sigma_p2": p["sigma_p2"],
                        "Pp": p["Pp"],
                        "Lp": p["Lp"],
                        "Mp": p["Mp"],
                    },
                    "water": {
                        "u_top_a": w["sigma_wa1"],
                        "u_bot_a": w["sigma_wa2"],
                        "u_top_p": w["sigma_wp1"],
                        "u_bot_p": w["sigma_wp2"],
                        "Mwa": w["Mwa"],
                        "Mwp": w["Mwp"],
                    },
                }
            )

        return {
            "driving_moment": driving,
            "resisting_moment": resisting,
            "sum_Ma": sum_ma,
            "sum_Mp": sum_mp,
            "sum_Mwa": sum_mwa,
            "sum_Mwp": sum_mwp,
            "sum_Mqh": sum_mqh,
            "active_rows": active_rows,
            "passive_rows": passive_rows,
            "water_rows": water_rows,
            "surcharge_rows": surcharge_rows,
            "details": details_legacy,
        }

    def _solve_fss(self) -> Tuple[float, Dict[str, Any]]:
        lo = 1.0
        hi = 3.0

        def f(fs_val: float) -> float:
            d = self._lateral_for_fs(fs_val)
            return d["resisting_moment"] - d["driving_moment"]

        flo = f(lo)
        fhi = f(hi)
        guard = 0
        while flo * fhi > 0 and guard < 12:
            hi += 1.0
            fhi = f(hi)
            guard += 1

        if flo * fhi > 0:
            # fallback: pick side with smaller |f|
            dlo = self._lateral_for_fs(lo)
            dhi = self._lateral_for_fs(hi)
            if abs(dlo["resisting_moment"] - dlo["driving_moment"]) <= abs(dhi["resisting_moment"] - dhi["driving_moment"]):
                return lo, dlo
            return hi, dhi

        best_data: Dict[str, Any] = {}
        for _ in range(60):
            mid = 0.5 * (lo + hi)
            dm = self._lateral_for_fs(mid)
            fm = dm["resisting_moment"] - dm["driving_moment"]
            best_data = dm
            if abs(fm) <= 1e-3 or abs(hi - lo) <= 1e-3:
                return mid, dm
            if flo * fm <= 0:
                hi = mid
                fhi = fm
            else:
                lo = mid
                flo = fm
        return 0.5 * (lo + hi), best_data

    def _build_weight_driving_table(self) -> Tuple[List[Dict[str, Any]], float, float, float]:
        b = max(0.1, self.dl - self.ds)
        lc = b / 2.0
        rows: List[Dict[str, Any]] = []
        sum_w = 0.0
        sum_mc = 0.0
        idx = 0
        for ly in self.layers:
            z1 = max(ly.top, self.ds)
            z2 = min(ly.bot, self.de)
            if z2 <= z1:
                continue
            idx += 1
            h = z2 - z1
            w = ly.gamma_t * h * b
            mc = w * lc
            rows.append(
                {
                    "No": idx,
                    "LayerID": ly.code,
                    "D1": _fmt_num(z1),
                    "D2": _fmt_num(z2),
                    "gamma_t": _fmt_num(ly.gamma_t),
                    "W": _fmt_num(w),
                    "Lc": _fmt_num(lc),
                    "Mc": _fmt_num(mc),
                }
            )
            sum_w += w
            sum_mc += mc
        return rows, sum_w, sum_mc, lc

    def _build_surcharge_driving_table(self, lc: float) -> Tuple[List[Dict[str, Any]], float]:
        rows: List[Dict[str, Any]] = []
        sum_mqv = 0.0
        b = max(0.1, self.dl - self.ds)

        if self.surcharge_mode == "sub" and self.sub_records:
            for i, rec in enumerate(self.sub_records, start=1):
                x1 = _f(rec.get("A", rec.get("x1", 0.0)))
                x2 = _f(rec.get("B", rec.get("x2", 0.0)))
                z = max(1e-6, abs(_f(rec.get("Z", rec.get("z", self.de)))))
                q = _f(rec.get("Q", rec.get("q", 0.0)))

                qv = self._surcharge_lateral_at(z, 1.0)
                pqv = qv * b
                mqv = pqv * lc
                sum_mqv += mqv
                rows.append(
                    {
                        "No": i,
                        "x1": _fmt_num(x1),
                        "x2": _fmt_num(x2),
                        "qv": _fmt_num(qv),
                        "Pqv": _fmt_num(pqv),
                        "Lqv": _fmt_num(lc),
                        "Mqv": _fmt_num(mqv),
                    }
                )
        else:
            qv = max(0.0, self.surcharge_q)
            pqv = qv * b
            mqv = pqv * lc
            sum_mqv += mqv
            rows.append(
                {
                    "No": 1,
                    "x1": _fmt_num(0.0),
                    "x2": _fmt_num(b),
                    "qv": _fmt_num(qv),
                    "Pqv": _fmt_num(pqv),
                    "Lqv": _fmt_num(lc),
                    "Mqv": _fmt_num(mqv),
                }
            )

        return rows, sum_mqv

    def _theta(self, z: float, side: str) -> float:
        b = max(0.1, self.dl - self.ds)
        if side == "active":
            x = max(0.0, self.de - z)
        else:
            x = max(0.0, z - self.de)
        return math.atan2(x, b)

    def _shear_side_tables(self, side: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], float]:
        upper: List[Dict[str, Any]] = []
        lower: List[Dict[str, Any]] = []
        total_v = 0.0

        b = max(0.1, self.dl - self.ds)
        z_lo = self.ds
        z_mid = self.de
        z_hi = self.dl

        # choose side depth range
        if side == "active":
            d_from, d_to = z_lo, z_mid
            k_side_name = "Ka"
        else:
            d_from, d_to = z_mid, z_hi
            k_side_name = "Kp"

        for i, ly in enumerate(self.layers, start=1):
            z1 = max(ly.top, d_from)
            z2 = min(ly.bot, d_to)
            if z2 <= z1:
                continue

            sig_v1 = self._vertical_total_sigma(z1)
            sig_v2 = self._vertical_total_sigma(z2)
            uo1, _ = self._u_out_in(ly, z1)
            uo2, _ = self._u_out_in(ly, z2)
            sig_v1_eff = sig_v1 - uo1
            sig_v2_eff = sig_v2 - uo2

            th1 = self._theta(z1, side)
            th2 = self._theta(z2, side)
            tstar = z2 - z1
            R = 1.0

            # use reduced strengths with FSs for consistency
            # if FSs not yet solved, use 1.0 fallback
            fs_s = max(1.0, _f(self.raw.get("_fs_solved", 1.0), 1.0))
            tan_phi_m = math.tan(math.radians(ly.phi_prime)) / fs_s if ly.drainage == "D" else 0.0
            phi_m = math.degrees(math.atan(tan_phi_m)) if ly.drainage == "D" else 0.0
            c_m = ly.c_prime / fs_s if ly.drainage == "D" else 0.0

            if side == "active":
                k_side = _active_k(phi_m if ly.drainage == "D" else 0.0, self.ka_method)
                k_c = math.sqrt(max(k_side, 0.0))
            else:
                k_side = _passive_k(phi_m if ly.drainage == "D" else 0.0, self.kp_method, self.delta_p_ratio)
                k_c = math.sqrt(max(k_side, 0.0))

            # Upper table values
            su1 = self._su_at(ly, sig_v1_eff) / fs_s if ly.drainage == "U" else 0.0
            su2 = self._su_at(ly, sig_v2_eff) / fs_s if ly.drainage == "U" else 0.0

            upper.append(
                {
                    "No": len(upper) + 1,
                    "LayerID": ly.code,
                    "D1": _fmt_num(z1),
                    "D2": _fmt_num(z2),
                    "Drainage": ly.drainage,
                    "gamma_t": _fmt_num(ly.gamma_t),
                    "c_prime": _fmt_num(c_m) if ly.drainage == "D" else "-",
                    "phi_prime": _fmt_num(phi_m) if ly.drainage == "D" else "-",
                    "Su1": _fmt_num(su1) if ly.drainage == "U" else "-",
                    "Su2": _fmt_num(su2) if ly.drainage == "U" else "-",
                    "theta1": _fmt_num(math.degrees(th1)),
                    "theta2": _fmt_num(math.degrees(th2)),
                    "sigma_v1": _fmt_num(sig_v1),
                    "sigma_v2": _fmt_num(sig_v2),
                    "sigma_v1_eff": _fmt_num(sig_v1_eff),
                    "sigma_v2_eff": _fmt_num(sig_v2_eff),
                    "K": _fmt_num(k_side, 4),
                }
            )

            # Lower table C/I/tau/V
            if ly.drainage == "U":
                K = 1.0
                sigma_vt = 0.5 * (sig_v1_eff + sig_v2_eff)
                c_star = 0.5 * (su1 + su2)
                C1 = c_star + 0.5 * (1.0 + K) * sigma_vt - 0.5 * (1.0 + K) * R * ly.gamma_t * tstar * math.sin(th1)
                C2 = 0.5 * (1.0 + K) * R * ly.gamma_t * tstar
                C3 = 0.5 * (1.0 - K) * sigma_vt + 0.5 * (1.0 - K) * R * ly.gamma_t * tstar * math.sin(th1)
                C4 = -0.5 * (1.0 - K) * R * ly.gamma_t * tstar
                tau1 = su1
                tau2 = su2
            else:
                K = k_side
                tan_phi = math.tan(math.radians(max(phi_m, 0.0)))
                gamma_eff = max(0.0, ly.gamma_t - GAMMA_W)
                if side == "active":
                    C1 = c_m + 0.5 * (1.0 + K) * sig_v1_eff * tan_phi - 0.5 * (1.0 + K) * R * gamma_eff * tstar * tan_phi * math.sin(th1) - c_m * k_c * tan_phi
                else:
                    C1 = c_m + 0.5 * (1.0 + K) * sig_v1_eff * tan_phi - 0.5 * (1.0 + K) * R * gamma_eff * tstar * tan_phi * math.sin(th1) + c_m * k_c * tan_phi
                C2 = 0.5 * (1.0 + K) * R * gamma_eff * tstar * tan_phi
                C3 = 0.5 * (1.0 - K) * sig_v1_eff * tan_phi
                C4 = -0.5 * (1.0 - K) * R * gamma_eff * tstar * tan_phi
                tau1 = max(0.0, c_m + sig_v1_eff * tan_phi)
                tau2 = max(0.0, c_m + sig_v2_eff * tan_phi)

            I1 = tstar
            I2 = (math.sin(th2) - math.sin(th1)) * b
            I3 = (math.cos(th1) - math.cos(th2)) * b
            I4 = (th2 - th1) * b
            V = C1 * I1 + C2 * I2 + C3 * I3 + C4 * I4
            V = max(0.0, V)
            total_v += V

            lower.append(
                {
                    "No": len(lower) + 1,
                    "LayerID": ly.code,
                    "C1": _fmt_num(C1),
                    "C2": _fmt_num(C2),
                    "C3": _fmt_num(C3),
                    "C4": _fmt_num(C4),
                    "I1": _fmt_num(I1),
                    "I2": _fmt_num(I2),
                    "I3": _fmt_num(I3),
                    "I4": _fmt_num(I4),
                    "tau1": _fmt_num(tau1),
                    "tau2": _fmt_num(tau2),
                    "V": _fmt_num(V),
                    "K_case": k_side_name,
                }
            )

        return upper, lower, total_v

    def _basal_heave(self) -> Dict[str, Any]:
        weight_rows, sum_w, sum_mc, lc = self._build_weight_driving_table()
        surcharge_rows, sum_mqv = self._build_surcharge_driving_table(lc)

        upper_a, lower_a, va = self._shear_side_tables("active")
        upper_p, lower_p, vp = self._shear_side_tables("passive")

        denom = max(1e-9, sum_mc + sum_mqv)
        width = max(0.1, self.dl - self.ds)
        fsh = ((va + vp) * width) / denom

        return {
            "factor_of_safety": round(fsh, 3),
            "status": "OK" if fsh >= self.fsh_req else "NG",
            "va": round(va, 3),
            "vp": round(vp, 3),
            "sum_mc": round(sum_mc, 3),
            "sum_mqv": round(sum_mqv, 3),
            "weight_rows": weight_rows,
            "sum_w": round(sum_w, 3),
            "surcharge_driving_rows": surcharge_rows,
            "shear_active_upper": upper_a,
            "shear_active_lower": lower_a,
            "shear_passive_upper": upper_p,
            "shear_passive_lower": lower_p,
        }

    def run(self) -> Dict[str, Any]:
        fs_s, lateral = self._solve_fss()
        self.raw["_fs_solved"] = fs_s
        heave = self._basal_heave()

        lateral_analysis = {
            "analysis_type": "Lateral force balance (strength reduction)",
            "factor_of_safety": round(fs_s, 3),
            "status": "OK" if fs_s >= self.fss_req else "NG",
            "required_fs": round(self.fss_req, 3),
            "driving_moment": round(lateral["driving_moment"], 3),
            "resisting_moment": round(lateral["resisting_moment"], 3),
            "sum_Ma": round(lateral["sum_Ma"], 3),
            "sum_Mp": round(lateral["sum_Mp"], 3),
            "sum_Mwa": round(lateral["sum_Mwa"], 3),
            "sum_Mwp": round(lateral["sum_Mwp"], 3),
            "sum_Mqh": round(lateral["sum_Mqh"], 3),
            "details": lateral["details"],
            # bundle used by excel generator
            "table_bundle": {
                "surcharge_lateral_rows": lateral["surcharge_rows"],
                "water_rows": lateral["water_rows"],
                "active_rows": lateral["active_rows"],
                "passive_rows": lateral["passive_rows"],
                "sum_Mqh": round(lateral["sum_Mqh"], 3),
                "sum_Mwa": round(lateral["sum_Mwa"], 3),
                "sum_Mwp": round(lateral["sum_Mwp"], 3),
                "sum_Ma": round(lateral["sum_Ma"], 3),
                "sum_Mp": round(lateral["sum_Mp"], 3),
                "driving_moment": round(lateral["driving_moment"], 3),
                "resisting_moment": round(lateral["resisting_moment"], 3),
            },
        }

        heave_analysis = {
            "analysis_type": "Basal heave check",
            "factor_of_safety": heave["factor_of_safety"],
            "required_fs": round(self.fsh_req, 3),
            "status": heave["status"],
            "va": heave["va"],
            "vp": heave["vp"],
            "sum_mc": heave["sum_mc"],
            "sum_mqv": heave["sum_mqv"],
            # bundle used by excel generator
            "table_bundle": {
                "weight_rows": heave["weight_rows"],
                "sum_w": heave["sum_w"],
                "sum_mc": heave["sum_mc"],
                "surcharge_driving_rows": heave["surcharge_driving_rows"],
                "sum_mqv": heave["sum_mqv"],
                "shear_active_upper": heave["shear_active_upper"],
                "shear_active_lower": heave["shear_active_lower"],
                "shear_passive_upper": heave["shear_passive_upper"],
                "shear_passive_lower": heave["shear_passive_lower"],
                "va": heave["va"],
                "vp": heave["vp"],
            },
        }

        tables = {
            **lateral_analysis.get("table_bundle", {}),
            **heave_analysis.get("table_bundle", {}),
        }

        return {
            "metadata": {
                "design_code": self.standard,
                "excavation_type": self.excavation_type,
                "clay_method": self.clay_method,
                "ka_method": self.ka_method,
                "kp_method": self.kp_method,
                "ds": self.ds,
                "de": self.de,
                "dl": self.dl,
                "ms": self.ms,
                "fss_r": self.fss_req,
                "fsh_r": self.fsh_req,
                "dl_de_ratio": round(self.dl / self.de, 4) if self.de > 0 else 0.0,
                "units": {
                    "length": "m",
                    "force": "tf",
                    "stress": "tf/m²",
                    "unit_weight": "tf/m³",
                    "moment": "tf-m/m",
                },
            },
            "lateral_analysis": lateral_analysis,
            "heave_analysis": heave_analysis,
            "tables": tables,
        }


def run_supported_tag_analysis(data: Dict[str, Any]) -> Dict[str, Any]:
    """Public API entrypoint used by Flask routes."""
    if not data:
        raise ValueError("Request data is required")

    engine = BracedClayWallEngine(data)
    if not engine.layers:
        raise ValueError("No valid soil layers were provided")
    if engine.de <= 0 or engine.dl <= 0:
        raise ValueError("Excavation depth De and wall length DL must be positive")
    if engine.dl <= engine.de:
        raise ValueError("Wall length DL must be greater than excavation depth De")

    return engine.run()


# ------------------------------
# Excel Export
# ------------------------------

def _header_style() -> Dict[str, Any]:
    thin = Side(style="thin", color="000000")
    return {
        "title_fill": PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"),
        "head_fill": PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid"),
        "title_font": Font(bold=True, color="FFFFFF", size=12),
        "head_font": Font(bold=True, color="000000", size=10),
        "sum_font": Font(bold=True, color="000000", size=10),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def _write_table(
    ws,
    title: str,
    headers: List[str],
    units: List[str],
    rows: List[Dict[str, Any]],
    keys: List[str],
    sum_row: List[Any] | None = None,
    start_row: int = 1,
    freeze: bool = True,
) -> int:
    st = _header_style()

    ws.cell(row=start_row, column=1, value=title)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
    t = ws.cell(row=start_row, column=1)
    t.fill = st["title_fill"]
    t.font = st["title_font"]
    t.alignment = Alignment(horizontal="center", vertical="center")

    r = start_row + 1
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = st["head_fill"]
        cell.font = st["head_font"]
        cell.border = st["border"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r += 1
    for c, u in enumerate(units, start=1):
        cell = ws.cell(row=r, column=c, value=u)
        cell.fill = st["head_fill"]
        cell.font = Font(size=9, italic=True)
        cell.border = st["border"]
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        r += 1
        for c, k in enumerate(keys, start=1):
            v = row.get(k, "-")
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = st["border"]
            cell.alignment = Alignment(horizontal="right" if isinstance(v, (int, float)) else "center", vertical="center")
            if isinstance(v, (int, float)):
                cell.number_format = "0.00"

    if sum_row is not None:
        r += 1
        for c, v in enumerate(sum_row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = st["border"]
            cell.font = st["sum_font"]
            cell.alignment = Alignment(horizontal="right" if isinstance(v, (int, float)) else "center", vertical="center")
            if isinstance(v, (int, float)):
                cell.number_format = "0.00"

    # simple width
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(10, min(26, len(headers[c - 1]) + 2))

    if freeze:
        ws.freeze_panes = f"A{start_row + 3}"

    return r + 1


def _symbols_rows() -> List[Tuple[str, str, str, str]]:
    return [
        ("Ds", "Deepest strut level depth", "m", "Geometry"),
        ("De", "Excavation depth", "m", "Geometry"),
        ("DL", "Wall length", "m", "Geometry"),
        ("Ms", "External applied moment per meter", "tf-m/m", "Input"),
        ("FSs", "Lateral force balance safety factor", "-", "Solved by iteration"),
        ("FSs,r", "Required lateral force balance safety factor", "-", "Input"),
        ("FSh", "Basal heave safety factor", "-", "Computed"),
        ("FSh,r", "Required basal heave safety factor", "-", "Input"),
        ("γt", "Total unit weight", "tf/m³", "Layer parameter"),
        ("γw", "Water unit weight", "tf/m³", "Assumed 1.0"),
        ("σv", "Vertical total stress", "tf/m²", "Depth-dependent"),
        ("σ'v", "Vertical effective stress", "tf/m²", "σv-u"),
        ("u", "Pore water pressure", "tf/m²", "Hydrostatic/seepage"),
        ("Ka", "Active pressure coefficient", "-", "Method-selected"),
        ("Kp", "Passive pressure coefficient", "-", "Method-selected"),
        ("Kahm", "Active horizontal coefficient (mobilized)", "-", "Table output"),
        ("Kachm", "Active cohesion coefficient (mobilized)", "-", "Table output"),
        ("Kphm", "Passive horizontal coefficient (mobilized)", "-", "Table output"),
        ("Kpchm", "Passive cohesion coefficient (mobilized)", "-", "Table output"),
        ("σa", "Active lateral stress", "tf/m²", "Endpoint value"),
        ("σp", "Passive lateral stress", "tf/m²", "Endpoint value"),
        ("Pa", "Active resultant force", "tf/m", "Segment resultant"),
        ("La", "Active lever arm", "m", "To datum"),
        ("Ma", "Active moment", "tf-m/m", "Pa·La"),
        ("Pp", "Passive resultant force", "tf/m", "Segment resultant"),
        ("Lp", "Passive lever arm", "m", "To datum"),
        ("Mp", "Passive moment", "tf-m/m", "Pp·Lp"),
        ("Pwa", "Outside water resultant force", "tf/m", "Driving"),
        ("Lwa", "Outside water lever arm", "m", "To datum"),
        ("Mwa", "Outside water moment", "tf-m/m", "Pwa·Lwa"),
        ("Pwp", "Inside water resultant force", "tf/m", "Resisting"),
        ("Lwp", "Inside water lever arm", "m", "To datum"),
        ("Mwp", "Inside water moment", "tf-m/m", "Pwp·Lwp"),
        ("Q", "Surcharge intensity", "tf/m²", "Input"),
        ("Sh", "Equivalent surcharge lateral pressure", "tf/m²", "SUC/manual"),
        ("Sh(z)", "Depth-dependent surcharge lateral pressure", "tf/m²", "SUB"),
        ("A", "SUB parameter A", "m", "Input"),
        ("B", "SUB parameter B", "m", "Input"),
        ("Z", "SUB depth variable", "m", "Input/eval"),
        ("Pqh", "Surcharge lateral resultant force", "tf/m", "Integrated"),
        ("Lqh", "Surcharge lateral lever arm", "m", "To datum"),
        ("Mqh", "Surcharge lateral moment", "tf-m/m", "Pqh·Lqh"),
        ("ΣMqh", "Sum of surcharge lateral moments", "tf-m/m", "Driving"),
        ("W", "Driving soil weight resultant", "tf/m", "Basal heave"),
        ("Lc", "Lever arm of driving soil weight", "m", "Basal heave"),
        ("Mc", "Driving soil weight moment", "tf-m/m", "W·Lc"),
        ("ΣMc", "Sum of driving soil moments", "tf-m/m", "Basal heave"),
        ("qv", "Equivalent surcharge vertical stress", "tf/m²", "Basal heave"),
        ("Pqv", "Surcharge vertical resultant force", "tf/m", "Basal heave"),
        ("Lqv", "Surcharge vertical lever arm", "m", "Basal heave"),
        ("Mqv", "Surcharge vertical moment", "tf-m/m", "Basal heave"),
        ("ΣMqv", "Sum of surcharge vertical moments", "tf-m/m", "Basal heave"),
        ("Va", "Shear resistance on retained side failure surface", "tf/m", "Integral-simplification"),
        ("Vp", "Shear resistance on excavation side failure surface", "tf/m", "Integral-simplification"),
        ("Vs", "Segment shear resistance contribution", "tf/m", "C/I expression"),
        ("C1..C4", "Integral simplification coefficients", "various", "Per segment"),
        ("I1..I4", "Integral simplification terms", "m", "Per segment"),
        ("τ", "Shear strength", "tf/m²", "Per segment"),
    ]


def generate_supported_tag_excel(project_info, lateral_result, heave_result, layers, input_data=None):
    """
    Generate Excel report with required multi-table sheet structure.
    Keeps existing route signature for backward compatibility.
    """
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Pull bundled tables from analysis results
    t_lat = (lateral_result or {}).get("table_bundle", {})
    t_heave = (heave_result or {}).get("table_bundle", {})

    ds = _f((input_data or {}).get("ds", 0.0))
    de = _f((input_data or {}).get("de", 0.0))
    dl = _f((input_data or {}).get("dl", 0.0))
    ms = _f((input_data or {}).get("ms", 0.0))

    # Single integrated data sheet (all tables)
    ws = wb.create_sheet("Data")
    next_row = _write_table(
        ws,
        "Inputs - Geometry",
        ["Parameter", "Symbol", "Value", "Unit"],
        ["", "", "", ""],
        [
            {"p": "Deepest strut level", "s": "Ds", "v": ds, "u": "m"},
            {"p": "Excavation depth", "s": "De", "v": de, "u": "m"},
            {"p": "Wall length", "s": "DL", "v": dl, "u": "m"},
            {"p": "External moment", "s": "Ms", "v": ms, "u": "tf-m/m"},
            {"p": "DL/De", "s": "DL/De", "v": (dl / de if de > 0 else 0.0), "u": "-"},
            {"p": "FSs required", "s": "FSs,r", "v": _f((lateral_result or {}).get("required_fs", 1.2)), "u": "-"},
            {"p": "FSh required", "s": "FSh,r", "v": _f((heave_result or {}).get("required_fs", 1.2)), "u": "-"},
        ],
        ["p", "s", "v", "u"],
        start_row=1,
        freeze=False,
    )

    # Inputs - Layers and Water
    layer_rows = []
    d0 = 0.0
    for i, ly in enumerate(layers or [], start=1):
        th = _f(ly.get("thickness", 0.0))
        if th <= 0 and _f(ly.get("depth", 0.0)) > d0:
            th = _f(ly.get("depth")) - d0
        d1, d2 = d0, d0 + max(0.0, th)
        d0 = d2
        layer_rows.append(
            {
                "No": i,
                "LayerID": ly.get("code", f"L{i}"),
                "Drainage": "U" if str(ly.get("type", "D")).upper() in ("U", "CLAY") else "D",
                "D1": round(d1, 2),
                "D2": round(d2, 2),
                "gamma_t": round(_f(ly.get("gamma", ly.get("gamma_t", 0.0))), 3),
                "c_prime": round(_f(ly.get("c", ly.get("c_prime", 0.0))), 3),
                "phi_prime": round(_f(ly.get("phi", ly.get("phi_prime", 0.0))), 3),
                "Su_or_ratio": round(_f(ly.get("su", ly.get("Su", 0.0))), 3),
                "SPT_N": ly.get("spt", ""),
                "Dw_exc": ly.get("dw_exc_layer", ly.get("dw_exc", "")),
                "Dw_ret": ly.get("dw_ret_layer", ly.get("dw_ret", "")),
                "Seepage": ly.get("seepage", "-"),
            }
        )
    next_row = _write_table(
        ws,
        "Inputs - Layers and Water",
        ["No", "Layer ID", "Drainage", "D1", "D2", "γt", "c'", "φ'", "Su/σ'v or Su", "SPT-N", "Dw,exc", "Dw,ret", "Seepage"],
        ["", "", "", "m", "m", "tf/m³", "tf/m²", "deg", "- / tf/m²", "", "m", "m", ""],
        layer_rows,
        ["No", "LayerID", "Drainage", "D1", "D2", "gamma_t", "c_prime", "phi_prime", "Su_or_ratio", "SPT_N", "Dw_exc", "Dw_ret", "Seepage"],
        start_row=next_row,
        freeze=False,
    )

    # Inputs - Surcharge
    next_row = _write_table(
        ws,
        "Inputs - Surcharge",
        ["Surcharge Mode", "Q", "Sh (Manual)", "SUB Record Count"],
        ["", "tf/m²", "tf/m²", ""],
        [
            {
                "mode": (project_info or {}).get("surcharge_mode", "manual"),
                "q": _f((project_info or {}).get("surcharge_q", 0.0)),
                "sh": _f((project_info or {}).get("surcharge_sh", 0.0)),
                "n": len((project_info or {}).get("sub_records", [])),
            }
        ],
        ["mode", "q", "sh", "n"],
        start_row=next_row,
        freeze=False,
    )

    # Surcharge-induced lateral pressure
    next_row = _write_table(
        ws,
        "Surcharge-induced Lateral Pressure",
        ["No", "Dqh1", "Dqh2", "σq1", "σq2", "Pqh", "Lqh", "Mqh"],
        ["", "m", "m", "tf/m²", "tf/m²", "tf/m", "m", "tf-m/m"],
        t_lat.get("surcharge_lateral_rows", []),
        ["No", "Dqh1", "Dqh2", "sigma_q1", "sigma_q2", "Pqh", "Lqh", "Mqh"],
        sum_row=["", "", "", "", "", "", "ΣMqh", _f(t_lat.get("sum_Mqh", 0.0))],
        start_row=next_row,
        freeze=False,
    )

    # Water pressure
    next_row = _write_table(
        ws,
        "Water Pressure",
        ["No", "Layer", "D1", "D2", "σwa1", "σwa2", "σwp1", "σwp2", "net1(wa-wp)", "net2(wa-wp)", "Pwa", "Lwa", "Mwa", "Pwp", "Lwp", "Mwp"],
        ["", "", "m", "m", "tf/m²", "tf/m²", "tf/m²", "tf/m²", "tf/m²", "tf/m²", "tf/m", "m", "tf-m/m", "tf/m", "m", "tf-m/m"],
        t_lat.get("water_rows", []),
        ["No", "LayerID", "D1", "D2", "sigma_wa1", "sigma_wa2", "sigma_wp1", "sigma_wp2", "net_1", "net_2", "Pwa", "Lwa", "Mwa", "Pwp", "Lwp", "Mwp"],
        sum_row=["", "", "", "", "", "", "", "", "", "", "", "ΣMwa", _f(t_lat.get("sum_Mwa", 0.0)), "", "ΣMwp", _f(t_lat.get("sum_Mwp", 0.0))],
        start_row=next_row,
        freeze=False,
    )

    # Active earth pressure
    next_row = _write_table(
        ws,
        "Active Earth Pressure (Retained Side)",
        ["No", "Layer", "D1", "D2", "Drainage", "γt", "c'm", "φ'm", "Su1", "Su2", "Kahm", "Kachm", "σv1", "σv2", "σ'v1", "σ'v2", "σa1", "σa2", "Pa", "La", "Ma"],
        ["", "", "m", "m", "", "tf/m³", "tf/m²", "deg", "tf/m²", "tf/m²", "", "", "tf/m²", "tf/m²", "tf/m²", "tf/m²", "tf/m²", "tf/m²", "tf/m", "m", "tf-m/m"],
        t_lat.get("active_rows", []),
        ["No", "LayerID", "D1", "D2", "Drainage", "gamma_t", "c_m", "phi_m", "Su1", "Su2", "Kahm", "Kachm", "sigma_v1", "sigma_v2", "sigma_v1_eff", "sigma_v2_eff", "sigma_a1", "sigma_a2", "Pa", "La", "Ma"],
        sum_row=["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "ΣMa", _f(t_lat.get("sum_Ma", 0.0))],
        start_row=next_row,
        freeze=False,
    )

    # Passive earth pressure
    next_row = _write_table(
        ws,
        "Passive Earth Pressure (Excavation Side)",
        ["No", "Layer", "D1", "D2", "Drainage", "γt", "c'm", "φ'm", "Su1", "Su2", "Kphm", "Kpchm", "σv1", "σv2", "σ'v1", "σ'v2", "σp1", "σp2", "Pp", "Lp", "Mp"],
        ["", "", "m", "m", "", "tf/m³", "tf/m²", "deg", "tf/m²", "tf/m²", "", "", "tf/m²", "tf/m²", "tf/m²", "tf/m²", "tf/m²", "tf/m²", "tf/m", "m", "tf-m/m"],
        t_lat.get("passive_rows", []),
        ["No", "LayerID", "D1", "D2", "Drainage", "gamma_t", "c_m", "phi_m", "Su1", "Su2", "Kphm", "Kpchm", "sigma_v1", "sigma_v2", "sigma_v1_eff", "sigma_v2_eff", "sigma_p1", "sigma_p2", "Pp", "Lp", "Mp"],
        sum_row=["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "ΣMp", _f(t_lat.get("sum_Mp", 0.0))],
        start_row=next_row,
        freeze=False,
    )

    # Basal heave - driving soil weight
    next_row = _write_table(
        ws,
        "Basal Heave - Driving Soil Weight",
        ["No", "Layer", "D1", "D2", "γt", "W", "Lc", "Mc"],
        ["", "", "m", "m", "tf/m³", "tf/m", "m", "tf-m/m"],
        t_heave.get("weight_rows", []),
        ["No", "LayerID", "D1", "D2", "gamma_t", "W", "Lc", "Mc"],
        sum_row=["", "", "", "", "ΣW", _f(t_heave.get("sum_w", 0.0)), "ΣMc", _f(t_heave.get("sum_mc", 0.0))],
        start_row=next_row,
        freeze=False,
    )

    # Basal heave - surcharge driving moment
    next_row = _write_table(
        ws,
        "Basal Heave - Surcharge Driving Moment",
        ["No", "x1", "x2", "qv", "Pqv", "Lqv", "Mqv"],
        ["", "m", "m", "tf/m²", "tf/m", "m", "tf-m/m"],
        t_heave.get("surcharge_driving_rows", []),
        ["No", "x1", "x2", "qv", "Pqv", "Lqv", "Mqv"],
        sum_row=["", "", "", "", "", "ΣMqv", _f(t_heave.get("sum_mqv", 0.0))],
        start_row=next_row,
        freeze=False,
    )

    # Basal heave - shear resistance (active side)
    r = _write_table(
        ws,
        "Basal Heave - Shear Resistance (Active Side) Upper",
        ["No", "Layer", "D1", "D2", "Drainage", "γt", "c'", "φ'", "Su1", "Su2", "θ1", "θ2", "σv1", "σv2", "σ'v1", "σ'v2", "K"],
        ["", "", "m", "m", "", "tf/m³", "tf/m²", "deg", "tf/m²", "tf/m²", "deg", "deg", "tf/m²", "tf/m²", "tf/m²", "tf/m²", "-"],
        t_heave.get("shear_active_upper", []),
        ["No", "LayerID", "D1", "D2", "Drainage", "gamma_t", "c_prime", "phi_prime", "Su1", "Su2", "theta1", "theta2", "sigma_v1", "sigma_v2", "sigma_v1_eff", "sigma_v2_eff", "K"],
        start_row=next_row,
        freeze=False,
    )
    next_row = _write_table(
        ws,
        "Basal Heave - Shear Resistance (Active Side) Lower",
        ["No", "Layer", "C1", "C2", "C3", "C4", "I1", "I2", "I3", "I4", "τ1", "τ2", "V"],
        ["", "", "", "", "", "", "m", "m", "m", "m", "tf/m²", "tf/m²", "tf/m"],
        t_heave.get("shear_active_lower", []),
        ["No", "LayerID", "C1", "C2", "C3", "C4", "I1", "I2", "I3", "I4", "tau1", "tau2", "V"],
        sum_row=["", "", "", "", "", "", "", "", "", "", "", "Va", _f(t_heave.get("va", 0.0))],
        start_row=r + 1,
        freeze=False,
    )

    # Basal heave - shear resistance (passive side)
    r = _write_table(
        ws,
        "Basal Heave - Shear Resistance (Passive Side) Upper",
        ["No", "Layer", "D1", "D2", "Drainage", "γt", "c'", "φ'", "Su1", "Su2", "θ1", "θ2", "σv1", "σv2", "σ'v1", "σ'v2", "K"],
        ["", "", "m", "m", "", "tf/m³", "tf/m²", "deg", "tf/m²", "tf/m²", "deg", "deg", "tf/m²", "tf/m²", "tf/m²", "tf/m²", "-"],
        t_heave.get("shear_passive_upper", []),
        ["No", "LayerID", "D1", "D2", "Drainage", "gamma_t", "c_prime", "phi_prime", "Su1", "Su2", "theta1", "theta2", "sigma_v1", "sigma_v2", "sigma_v1_eff", "sigma_v2_eff", "K"],
        start_row=next_row,
        freeze=False,
    )
    next_row = _write_table(
        ws,
        "Basal Heave - Shear Resistance (Passive Side) Lower",
        ["No", "Layer", "C1", "C2", "C3", "C4", "I1", "I2", "I3", "I4", "τ1", "τ2", "V"],
        ["", "", "", "", "", "", "m", "m", "m", "m", "tf/m²", "tf/m²", "tf/m"],
        t_heave.get("shear_passive_lower", []),
        ["No", "LayerID", "C1", "C2", "C3", "C4", "I1", "I2", "I3", "I4", "tau1", "tau2", "V"],
        sum_row=["", "", "", "", "", "", "", "", "", "", "", "Vp", _f(t_heave.get("vp", 0.0))],
        start_row=r + 1,
        freeze=False,
    )

    # Summary
    fs_s = _f(lateral_result.get("factor_of_safety", 0.0)) if lateral_result else 0.0
    fs_h = _f(heave_result.get("factor_of_safety", 0.0)) if heave_result else 0.0
    fs_s_req = _f(lateral_result.get("required_fs", 1.2)) if lateral_result else 1.2
    fs_h_req = _f(heave_result.get("required_fs", 1.2)) if heave_result else 1.2
    _write_table(
        ws,
        "Summary",
        ["Check", "Computed", "Required", "Status"],
        ["", "-", "-", ""],
        [
            {
                "n": "Lateral force balance FS_s",
                "c": round(fs_s, 2),
                "r": round(fs_s_req, 2),
                "s": "OK" if fs_s >= fs_s_req else "NG",
            },
            {
                "n": "Basal heave FS_h",
                "c": round(fs_h, 2),
                "r": round(fs_h_req, 2),
                "s": "OK" if fs_h >= fs_h_req else "NG",
            },
            {
                "n": "Driving moment (ΣMa + ΣMwa + ΣMqh)",
                "c": _f(t_lat.get("driving_moment", 0.0)),
                "r": "-",
                "s": "tf-m/m",
            },
            {
                "n": "Resisting moment (ΣMp + ΣMwp + Ms)",
                "c": _f(t_lat.get("resisting_moment", 0.0)),
                "r": "-",
                "s": "tf-m/m",
            },
            {
                "n": "Basal heave formula ((Va+Vp)(DL-Ds)/(ΣMc+ΣMqv))",
                "c": round(fs_h, 2),
                "r": round(fs_h_req, 2),
                "s": "OK" if fs_h >= fs_h_req else "NG",
            },
        ],
        ["n", "c", "r", "s"],
        start_row=next_row,
        freeze=False,
    )
    ws.freeze_panes = "A4"

    # Separate symbols sheet
    ws = wb.create_sheet("Symbol Description")
    _write_table(
        ws,
        "Symbol Description",
        ["Symbol", "Description (EN)", "Unit", "Notes"],
        ["", "", "", ""],
        [{"a": a, "b": b, "c": c, "d": d} for (a, b, c, d) in _symbols_rows()],
        ["a", "b", "c", "d"],
        start_row=1,
    )

    # fixed number format tuning per sheet
    for sh in wb.worksheets:
        sh.sheet_view.showGridLines = False

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
