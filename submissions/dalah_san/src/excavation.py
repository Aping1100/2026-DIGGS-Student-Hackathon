"""
Deep Excavation Analysis Module
Contains uplift analysis, sand boil analysis, and Excel output.
Supports imperial and metric units.
"""

import io
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as patches

# Unit conversion constants
FT_TO_M = 0.3048
M_TO_FT = 1.0 / FT_TO_M
PCF_TO_TF_M3 = 0.016018
TF_M3_TO_PCF = 1.0 / PCF_TO_TF_M3
TF_M2_TO_PSF = 204.816
PSF_TO_TF_M2 = 1.0 / TF_M2_TO_PSF
KN_PER_TF = 9.80665
PCF_PER_KN_M3 = 6.36588
KPA_PER_TF_M2 = KN_PER_TF


def _convert_inputs_to_metric(data, unit_system):
    """Normalize user input into internal metric tf-based units for calculation."""
    def _safe_int(v, default=0):
        try:
            return int(v)
        except Exception:
            return default

    if unit_system == 'imperial':
        layers = []
        for L in data.get('layers', []):
            layers.append({
                'bot_depth': float(L.get('bot_depth', 0)) * FT_TO_M,
                'gamma': float(L.get('gamma', L.get('unit_weight', 0))) * PCF_TO_TF_M3,
                'type': L.get('type', 'D'),
                'code': L.get('code', '')
            })
        stages = []
        for s in data.get('stages', []):
            wls = s.get('water_levels', [s.get('water_level', data.get('gwt_gl', 2.0))])
            if not isinstance(wls, list):
                wls = [wls]
            stages.append({
                'name': s.get('name', 'Stage'),
                'depth': float(s.get('depth', 0)) * FT_TO_M,
                'water_levels': [float(w) * FT_TO_M for w in wls]
            })
        return {
            'wall_length': float(data.get('wall_length', 21.0)) * FT_TO_M,
            'gwt_gl': float(data.get('gwt_gl', 2.0)) * FT_TO_M,
            'interface_depth': float(data.get('interface_depth', 4.0)) * FT_TO_M,
            'interface_desc': data.get('interface_desc', 'Interface'),
            'interface_index': _safe_int(data.get('interface_index', 0), 0),
            'layers': layers,
            'stages': stages,
            'analyze_uplift': data.get('analyze_uplift', True),
            'analyze_sand_boil': data.get('analyze_sand_boil', True),
            'fs_u_required': float(data.get('fs_u_required', 1.2)),
            'fs_p1_required': float(data.get('fs_p1_required', 1.5)),
            'fs_p2_required': float(data.get('fs_p2_required', 2.0)),
        }

    layers = []
    for L in data.get('layers', []):
        layers.append({
            'bot_depth': float(L.get('bot_depth', 0)),
            'gamma': float(L.get('gamma', L.get('unit_weight', 0))) / KN_PER_TF,  # kN/m³ -> tf/m³
            'type': L.get('type', 'D'),
            'code': L.get('code', '')
        })
    stages = []
    for s in data.get('stages', []):
        wls = s.get('water_levels', [s.get('water_level', data.get('gwt_gl', 2.0))])
        if not isinstance(wls, list):
            wls = [wls]
        stages.append({
            'name': s.get('name', 'Stage'),
            'depth': float(s.get('depth', 0)),
            'water_levels': [float(w) for w in wls]
        })
    return {
        'wall_length': float(data.get('wall_length', 21.0)),
        'gwt_gl': float(data.get('gwt_gl', 2.0)),
        'interface_depth': float(data.get('interface_depth', 4.0)),
        'interface_desc': data.get('interface_desc', 'Interface'),
        'interface_index': _safe_int(data.get('interface_index', 0), 0),
        'layers': layers,
        'stages': stages,
        'analyze_uplift': data.get('analyze_uplift', True),
        'analyze_sand_boil': data.get('analyze_sand_boil', True),
        'fs_u_required': float(data.get('fs_u_required', 1.2)),
        'fs_p1_required': float(data.get('fs_p1_required', 1.5)),
        'fs_p2_required': float(data.get('fs_p2_required', 2.0)),
    }


def _resolve_interface(layers, interface_depth, interface_desc, interface_index):
    """Resolve target interface from user input and U-layers.

    Interface must align to a U-layer bottom depth to keep stage water-level indexing consistent.
    """
    interfaces = []
    for layer in layers:
        try:
            layer_type = str(layer.get('type', 'D')).strip().upper()
            if layer_type in ('U', 'UNDRAINED'):
                interfaces.append({
                    'depth': float(layer.get('bot_depth', 0.0)),
                    'code': str(layer.get('code', 'U') or 'U')
                })
        except Exception:
            continue

    if not interfaces:
        raise ValueError("No U-layer found. At least one U (undrained) layer is required as interface.")

    # Priority: explicit interface_index -> nearest by interface_depth -> first U-layer
    idx = 0
    try:
        idx_candidate = int(interface_index)
        if 0 <= idx_candidate < len(interfaces):
            idx = idx_candidate
        else:
            raise ValueError()
    except Exception:
        try:
            target_depth = float(interface_depth)
            idx = min(range(len(interfaces)), key=lambda i: abs(interfaces[i]['depth'] - target_depth))
        except Exception:
            idx = 0

    sel = interfaces[idx]
    code = str(interface_desc or '').strip() or str(sel.get('code') or f"Interface {idx + 1}")
    return {
        'index': idx,
        'depth': float(sel['depth']),
        'code': code,
        'interfaces': interfaces
    }


def _convert_outputs_to_display(result, unit_system):
    """Convert internal tf-based outputs to selected display units for display/Excel."""
    if unit_system == 'metric':
        for ur in result.get('uplift_results', []):
            for st in ur.get('stages', []):
                st['Uw'] = st['Uw'] * KPA_PER_TF_M2 if isinstance(st.get('Uw'), (int, float)) else st['Uw']
                st['Weight'] = st['Weight'] * KPA_PER_TF_M2 if isinstance(st.get('Weight'), (int, float)) else st['Weight']
        for sr in result.get('sand_boil_results', []):
            sr['gamma_sub'] = sr['gamma_sub'] * KN_PER_TF if isinstance(sr.get('gamma_sub'), (int, float)) else sr['gamma_sub']
            sr['gamma_sub_D'] = sr['gamma_sub_D'] * KPA_PER_TF_M2 if isinstance(sr.get('gamma_sub_D'), (int, float)) else sr['gamma_sub_D']
            sr['gamma_sub_delta_hw'] = sr['gamma_sub_delta_hw'] * KPA_PER_TF_M2 if isinstance(sr.get('gamma_sub_delta_hw'), (int, float)) else sr['gamma_sub_delta_hw']
            sr['gamma_w_delta_hw'] = sr['gamma_w_delta_hw'] * KPA_PER_TF_M2 if isinstance(sr.get('gamma_w_delta_hw'), (int, float)) else sr['gamma_w_delta_hw']
        return result

    if unit_system != 'imperial':
        return result

    m = result.get('metadata', {})
    m['wall_length'] = m.get('wall_length', 0) * M_TO_FT
    m['gwt_gl'] = m.get('gwt_gl', 0) * M_TO_FT
    m['interface_depth'] = m.get('interface_depth', 0) * M_TO_FT
    for ur in result.get('uplift_results', []):
        ur['interface_depth'] = ur.get('interface_depth', 0) * M_TO_FT
        ur['interface'] = f"GL-{ur['interface_depth']:.2f}ft ({ur.get('interface_code', '')})"
        for st in ur.get('stages', []):
            st['De'] = st['De'] * M_TO_FT if isinstance(st.get('De'), (int, float)) else st['De']
            st['Dw'] = st['Dw'] * M_TO_FT if isinstance(st.get('Dw'), (int, float)) else st['Dw']
            st['Uw'] = st['Uw'] * TF_M2_TO_PSF if isinstance(st.get('Uw'), (int, float)) else st['Uw']
            st['Weight'] = st['Weight'] * TF_M2_TO_PSF if isinstance(st.get('Weight'), (int, float)) else st['Weight']
    for sr in result.get('sand_boil_results', []):
        sr['De'] = sr['De'] * M_TO_FT if isinstance(sr.get('De'), (int, float)) else sr['De']
        sr['dHw'] = sr['dHw'] * M_TO_FT if isinstance(sr.get('dHw'), (int, float)) else sr['dHw']
        sr['D'] = sr['D'] * M_TO_FT if isinstance(sr.get('D'), (int, float)) else sr['D']
        sr['gamma_sub'] = sr['gamma_sub'] * TF_M3_TO_PCF if isinstance(sr.get('gamma_sub'), (int, float)) else sr['gamma_sub']
        sr['gamma_sub_D'] = sr['gamma_sub_D'] * TF_M2_TO_PSF if isinstance(sr.get('gamma_sub_D'), (int, float)) else sr['gamma_sub_D']
        sr['gamma_sub_delta_hw'] = sr['gamma_sub_delta_hw'] * TF_M2_TO_PSF if isinstance(sr.get('gamma_sub_delta_hw'), (int, float)) else sr['gamma_sub_delta_hw']
        sr['gamma_w_delta_hw'] = sr['gamma_w_delta_hw'] * TF_M2_TO_PSF if isinstance(sr.get('gamma_w_delta_hw'), (int, float)) else sr['gamma_w_delta_hw']
    return result


def run_excavation_analysis(data):
    """
    Run excavation analysis (uplift and sand boil).
    Supports unit_system: 'metric' | 'imperial'
    """
    try:
        unit_system = data.get('unit_system', 'metric')
        calc_data = _convert_inputs_to_metric(data, unit_system)

        wall_length = float(calc_data.get('wall_length', 21.0))
        gwt_gl = float(calc_data.get('gwt_gl', 2.0))
        interface_depth = float(calc_data.get('interface_depth', 4.0))
        interface_desc = calc_data.get('interface_desc', 'Interface')
        interface_index = int(calc_data.get('interface_index', 0))
        layers = calc_data.get('layers', [])
        stages = calc_data.get('stages', [])
        analyze_uplift = calc_data.get('analyze_uplift', True)
        analyze_sand_boil = calc_data.get('analyze_sand_boil', True)
        fs_u_required = float(calc_data.get('fs_u_required', 1.2))
        fs_p1_required = float(calc_data.get('fs_p1_required', 1.5))
        fs_p2_required = float(calc_data.get('fs_p2_required', 2.0))
        selected_interface = _resolve_interface(layers, interface_depth, interface_desc, interface_index)

        metadata = {
            'wall_length': wall_length,
            'gwt_gl': gwt_gl,
            'interface_depth': float(selected_interface['depth']),
            'interface_desc': selected_interface['code'],
            'interface_index': int(selected_interface['index']),
            'fs_u_required': fs_u_required,
            'fs_p1_required': fs_p1_required,
            'fs_p2_required': fs_p2_required,
            'unit_system': unit_system
        }

        uplift_results = []
        if analyze_uplift:
            uplift_data = calculate_uplift(layers, stages, gwt_gl, selected_interface, fs_u_required)
            if uplift_data and uplift_data.get('stages'):
                interface_groups = {}
                for stage_result in uplift_data['stages']:
                    ifd = stage_result.get('interface_depth', 0)
                    ifc = stage_result.get('interface_code', '')
                    key = f"{ifd:.2f}_{ifc}"
                    if key not in interface_groups:
                        interface_groups[key] = {'interface_depth': ifd, 'interface_code': ifc, 'stages': []}
                    if stage_result.get('total_weight') == '-':
                        interface_groups[key]['stages'].append({
                            'Stage': stage_result.get('stage', ''),
                            'De': stage_result.get('excavation_depth', 0),
                            'Dw': stage_result.get('water_level', gwt_gl),
                            'Uw': '-', 'Weight': '-', 'FSu': '-', 'Status': 'OK',
                            'Result': 'Impermeable layer above interface has been excavated, no check required !',
                            'Note': 'Impermeable layer above interface has been excavated, no check required !'
                        })
                    else:
                        interface_groups[key]['stages'].append({
                            'Stage': stage_result.get('stage', ''),
                            'De': stage_result.get('excavation_depth', 0),
                            'Dw': stage_result.get('water_level', gwt_gl),
                            'Uw': stage_result.get('water_pressure', 0) if stage_result.get('water_pressure') != '-' else '-',
                            'Weight': stage_result.get('total_weight', 0) if stage_result.get('total_weight') != '-' else '-',
                            'FSu': stage_result.get('fs_u', 0) if stage_result.get('fs_u') != '-' else '-',
                            'Status': 'OK' if stage_result.get('is_safe', False) else 'NG',
                            'Result': stage_result.get('status', ''),
                            'Note': stage_result.get('status', '')
                        })
                for k, v in interface_groups.items():
                    uplift_results.append({
                        'interface': f"GL-{v['interface_depth']:.2f}m ({v['interface_code']})",
                        'interface_depth': v['interface_depth'],
                        'interface_code': v['interface_code'],
                        'stages': v['stages']
                    })

        sand_boil_results = []
        if analyze_sand_boil and stages:
            sand_boil_data = calculate_sand_boil(
                layers, list(stages), gwt_gl, selected_interface, wall_length,
                fs_p1_required, fs_p2_required
            )
            sb_stages = sand_boil_data.get('stages', []) if sand_boil_data else []
            if sb_stages:
                for stage_result in sb_stages:
                    if 'error' in stage_result:
                        d_val = stage_result.get('D', '-')
                        d_val = d_val if isinstance(d_val, (int, float)) else '-'
                        sand_boil_results.append({
                            'Stage': stage_result.get('stage', ''),
                            'De': stage_result.get('excavation_depth', 0),
                            'dHw': stage_result.get('delta_h_w', '-') if isinstance(stage_result.get('delta_h_w'), (int, float)) else '-',
                            'D': d_val, 'gamma_sub': '-', 'gamma_sub_D': '-',
                            'gamma_sub_delta_hw': '-', 'gamma_w_delta_hw': '-',
                            'FSp1': '-', 'FSp2': '-', 'Status': 'NG',
                            'Result': stage_result.get('error', 'Error'),
                            'Note': stage_result.get('error', 'Error')
                        })
                    else:
                        sand_boil_results.append({
                            'Stage': stage_result.get('stage', ''),
                            'De': stage_result.get('excavation_depth', 0),
                            'dHw': stage_result.get('delta_h_w', 0),
                            'D': stage_result.get('D', 0) if stage_result.get('D') != '-' else '-',
                            'gamma_sub': round(float(stage_result.get('gamma_sub', 0)), 3) if stage_result.get('gamma_sub') != '-' else '-',
                            'gamma_sub_D': round(float(stage_result.get('gamma_sub_D', 0)), 3) if stage_result.get('gamma_sub_D') != '-' else '-',
                            'gamma_sub_delta_hw': round(float(stage_result.get('gamma_sub_delta_hw', 0)), 3) if stage_result.get('gamma_sub_delta_hw') != '-' else '-',
                            'gamma_w_delta_hw': round(float(stage_result.get('gamma_w_delta_hw', 0)), 3) if stage_result.get('gamma_w_delta_hw') != '-' else '-',
                            'FSp1': round(float(stage_result.get('fs_p1', 0)), 2) if stage_result.get('fs_p1') != '-' else '-',
                            'FSp2': round(float(stage_result.get('fs_p2', 0)), 2) if stage_result.get('fs_p2') != '-' else '-',
                            'Status': 'OK' if stage_result.get('is_safe', False) else 'NG',
                            'Result': stage_result.get('status', ''),
                            'Note': stage_result.get('status', '')
                        })

        result = {
            'metadata': metadata,
            'uplift_results': uplift_results,
            'sand_boil_results': sand_boil_results
        }
        return _convert_outputs_to_display(result, unit_system)
    except Exception as e:
        raise ValueError(f"Analysis calculation error: {str(e)}")


def calculate_uplift(layers, stages, gwt_gl, selected_interface, fs_u_required):
    """Calculate uplift factor of safety for one selected interface."""
    results = []
    interface_idx = int(selected_interface.get('index', 0))
    interface_depth_val = float(selected_interface.get('depth', 0.0))
    interface_code = str(selected_interface.get('code', 'Interface'))

    for stage in stages:
        excavation_depth = float(stage.get('depth', 0))
        water_levels = stage.get('water_levels', [stage.get('water_level', gwt_gl)])
        if water_levels and len(water_levels) > interface_idx:
            stage_water_level = water_levels[interface_idx]
        elif water_levels:
            stage_water_level = water_levels[-1]
        else:
            stage_water_level = gwt_gl

        if (excavation_depth < 0 and excavation_depth <= interface_depth_val) or \
           (excavation_depth >= 0 and excavation_depth >= interface_depth_val):
            results.append({
                'stage': stage.get('name', ''),
                'interface_depth': interface_depth_val,
                'interface_code': interface_code,
                'excavation_depth': excavation_depth,
                'total_weight': '-', 'water_pressure': '-', 'fs_u': '-',
                'fs_u_required': fs_u_required, 'is_safe': True,
                'status': 'Impermeable layer above interface has been excavated, no check required !',
                'water_level': stage_water_level
            })
            continue

        total_weight = 0.0
        pd = excavation_depth
        for L in layers:
            bd = float(L.get('bot_depth', 0))
            gamma = float(L.get('gamma', L.get('unit_weight', 0)))
            lt = pd
            if (excavation_depth < 0 and bd >= excavation_depth) or \
               (excavation_depth >= 0 and bd <= excavation_depth):
                pd = bd
                continue
            if (excavation_depth < 0 and lt >= interface_depth_val) or \
               (excavation_depth >= 0 and lt >= interface_depth_val):
                break
            top_r = max(lt, excavation_depth)
            bot_r = min(bd, interface_depth_val)
            thick = abs(bot_r - top_r)
            if thick > 0 and gamma > 0:
                total_weight += gamma * thick
            pd = bd

        water_pressure = 1.0 * abs(stage_water_level - interface_depth_val)
        fs_u = total_weight / water_pressure if water_pressure > 0 else 999.0
        # Avoid floating-point error: FS >= required passes; compare at 2-decimal precision
        is_safe = round(fs_u, 2) >= round(float(fs_u_required), 2)
        results.append({
            'stage': stage.get('name', ''),
            'interface_depth': interface_depth_val,
            'interface_code': interface_code,
            'excavation_depth': excavation_depth,
            'total_weight': round(total_weight, 3),
            'water_pressure': round(water_pressure, 3),
            'fs_u': round(fs_u, 2),
            'fs_u_required': fs_u_required,
            'is_safe': is_safe,
            'status': f'FSu={fs_u:.2f} {"OK" if is_safe else "NG"}',
            'water_level': stage_water_level
        })
    return {'stages': results}


def calculate_sand_boil(layers, stages, gwt_gl, selected_interface, wall_length, fs_p1_required, fs_p2_required):
    """
    Calculate sand boil factor of safety.
    D = distance from excavation surface to retaining wall bottom (embedment depth) = |wall_length - excavation_depth|
    D is computed independently per stage.
    """
    results = []
    interface_idx = int(selected_interface.get('index', 0))
    interface_depth = float(selected_interface.get('depth', 0.0))
    interface_code = str(selected_interface.get('code', 'Interface'))
    for stage in stages:
        stage_name = stage.get('name', 'Stage')
        excavation_depth = float(stage.get('depth', 0))
        water_levels = stage.get('water_levels', [stage.get('water_level', gwt_gl)])
        if water_levels and len(water_levels) > interface_idx:
            stage_water_level = water_levels[interface_idx]
        elif water_levels:
            stage_water_level = water_levels[-1]
        else:
            stage_water_level = gwt_gl

        has_cohesive_between = False
        prev_depth = 0
        check_top = excavation_depth
        check_bottom = wall_length
        for layer in layers:
            bot_depth = float(layer.get('bot_depth', 0))
            layer_type = str(layer.get('type', 'D')).strip().upper()
            layer_top = prev_depth
            if layer_type in ('U', 'UNDRAINED'):
                if excavation_depth < 0:
                    if layer_top >= check_bottom and bot_depth <= check_top:
                        has_cohesive_between = True
                        break
                else:
                    if layer_top < check_bottom and bot_depth > check_top:
                        has_cohesive_between = True
                        break
            prev_depth = bot_depth

        # D = distance from excavation surface to retaining wall bottom; computed per stage
        D_val = abs(wall_length - excavation_depth)
        D_val = round(D_val, 2) if D_val > 0 else 0

        if has_cohesive_between:
            results.append({
                'stage': stage_name, 'excavation_depth': round(excavation_depth, 2),
                'delta_h_w': round(stage_water_level - excavation_depth, 2),
                'D': D_val, 'gamma_sub': '-', 'gamma_sub_D': '-', 'gamma_sub_delta_hw': '-',
                'gamma_w_delta_hw': '-', 'fs_p1': '-', 'fs_p2': '-',
                'fs_p1_required': fs_p1_required, 'fs_p2_required': fs_p2_required,
                'is_safe_p1': True, 'is_safe_p2': True, 'is_safe': True,
                'status': 'Cohesive soil layer exists below excavation surface, no check required!',
                'water_level': round(stage_water_level, 2),
                'interface_depth': interface_depth,
                'interface_code': interface_code
            })
            continue

        excavation_layer = None
        prev_depth = 0
        for layer in layers:
            bot_depth = float(layer.get('bot_depth', 0))
            layer_type = str(layer.get('type', 'D')).strip().upper()
            layer_top = prev_depth
            if excavation_depth < 0:
                if layer_top >= excavation_depth >= bot_depth:
                    excavation_layer = {'type': layer_type, 'top': layer_top, 'bottom': bot_depth}
                    break
            else:
                if layer_top <= excavation_depth <= bot_depth:
                    excavation_layer = {'type': layer_type, 'top': layer_top, 'bottom': bot_depth}
                    break
            prev_depth = bot_depth

        if excavation_layer and excavation_layer['type'] in ('U', 'UNDRAINED'):
            results.append({
                'stage': stage_name, 'excavation_depth': round(excavation_depth, 2),
                'delta_h_w': round(stage_water_level - excavation_depth, 2),
                'D': D_val, 'gamma_sub': '-', 'gamma_sub_D': '-', 'gamma_sub_delta_hw': '-',
                'gamma_w_delta_hw': '-', 'fs_p1': '-', 'fs_p2': '-',
                'fs_p1_required': fs_p1_required, 'fs_p2_required': fs_p2_required,
                'is_safe_p1': True, 'is_safe_p2': True, 'is_safe': True,
                'status': 'Cohesive soil layer exists below excavation surface, no check required!',
                'water_level': round(stage_water_level, 2),
                'interface_depth': interface_depth,
                'interface_code': interface_code
            })
            continue

        D = D_val
        if D <= 0:
            results.append({
                'stage': stage_name, 'excavation_depth': round(excavation_depth, 2),
                'D': D_val, 'delta_h_w': round(stage_water_level - excavation_depth, 2),
                'error': 'Wall bottom must be below excavation surface (D > 0 required)',
                'interface_depth': interface_depth,
                'interface_code': interface_code
            })
            continue

        delta_h_w = abs(stage_water_level - excavation_depth)
        gamma_sub_weighted_sum = 0.0
        total_thickness = 0.0
        prev_depth = 0
        calc_range_top = excavation_depth
        calc_range_bottom = wall_length

        for layer in layers:
            bot_depth = float(layer.get('bot_depth', 0))
            gamma = float(layer.get('gamma', layer.get('unit_weight', 0)))
            layer_type = str(layer.get('type', 'D')).strip().upper()
            layer_top = prev_depth
            if layer_type in ('U', 'UNDRAINED'):
                prev_depth = bot_depth
                if (excavation_depth < 0 and prev_depth <= calc_range_bottom) or \
                   (excavation_depth >= 0 and prev_depth >= calc_range_bottom):
                    break
                continue
            if excavation_depth < 0:
                if bot_depth >= excavation_depth:
                    prev_depth = bot_depth
                    continue
                if layer_top >= calc_range_bottom:
                    break
            else:
                if bot_depth <= excavation_depth:
                    prev_depth = bot_depth
                    continue
                if layer_top >= calc_range_bottom:
                    break
            if calc_range_top < 0 and calc_range_bottom < 0:
                layer_top_in_range = max(layer_top, calc_range_top)
                layer_bottom_in_range = min(bot_depth, calc_range_bottom)
            else:
                layer_top_in_range = max(layer_top, calc_range_top)
                layer_bottom_in_range = min(bot_depth, calc_range_bottom)
            thickness = abs(layer_bottom_in_range - layer_top_in_range)
            if thickness > 0 and gamma > 0:
                gamma_sub = max(gamma - 1.0, 0)
                gamma_sub_weighted_sum += gamma_sub * thickness
                total_thickness += thickness
            prev_depth = bot_depth
            if (excavation_depth < 0 and prev_depth <= calc_range_bottom) or \
               (excavation_depth >= 0 and prev_depth >= calc_range_bottom):
                break

        gamma_sub_avg = gamma_sub_weighted_sum / total_thickness if total_thickness > 0 else 0.9
        gamma_sub_D = gamma_sub_avg * D
        gamma_sub_delta_hw = gamma_sub_avg * delta_h_w
        gamma_w_delta_hw = 1.0 * delta_h_w
        fs_p1 = (2 * gamma_sub_D) / gamma_w_delta_hw if gamma_w_delta_hw > 0 else 999.0
        fs_p2 = (gamma_sub_delta_hw + 2 * gamma_sub_D) / gamma_w_delta_hw if gamma_w_delta_hw > 0 else 999.0
        # Avoid floating-point error: FS >= required passes; compare at 2-decimal precision
        is_safe_p1 = round(fs_p1, 2) >= round(float(fs_p1_required), 2)
        is_safe_p2 = round(fs_p2, 2) >= round(float(fs_p2_required), 2)
        is_safe = is_safe_p1 and is_safe_p2
        status = f'FSp1≥{fs_p1_required:.2f}, FSp2≥{fs_p2_required:.2f}, OK!' if is_safe else \
                 (f'FSp1<{fs_p1_required:.2f}' if not is_safe_p1 else f'FSp2<{fs_p2_required:.2f}') + ', NG!'

        results.append({
            'stage': stage_name, 'excavation_depth': round(excavation_depth, 2),
            'D': round(D, 2), 'delta_h_w': round(delta_h_w, 2),
            'gamma_sub': round(gamma_sub_avg, 3), 'gamma_sub_D': round(gamma_sub_D, 3),
            'gamma_sub_delta_hw': round(gamma_sub_delta_hw, 3),
            'gamma_w_delta_hw': round(gamma_w_delta_hw, 3),
            'fs_p1': round(fs_p1, 2), 'fs_p2': round(fs_p2, 2),
            'fs_p1_required': fs_p1_required, 'fs_p2_required': fs_p2_required,
            'is_safe_p1': is_safe_p1, 'is_safe_p2': is_safe_p2, 'is_safe': is_safe,
            'status': status, 'water_level': round(stage_water_level, 2),
            'interface_depth': interface_depth,
            'interface_code': interface_code
        })

    return {'stages': results, 'fs_p1_required': fs_p1_required, 'fs_p2_required': fs_p2_required}


def draw_excavation_profile(layers, wall_length, excavation_depth, water_levels, interface_depth, stage_name, max_depth=None):
    """
    Draw excavation profile.

    Parameters:
        layers: List of soil layers
        wall_length: Retaining wall length (m)
        excavation_depth: Excavation depth (m)
        water_levels: List of water levels (m)
        interface_depth: Impermeable layer depth (m)
        stage_name: Stage name
        max_depth: Maximum depth (for uniform y-axis)

    Returns:
        BytesIO object containing image data
    """
    fig, ax = plt.subplots(figsize=(8, 10))
    
    if max_depth is None:
        max_depth = max([layer.get('bot_depth', 0) for layer in layers] + [wall_length, excavation_depth]) + 5
    
    ax.set_xlim(0, 10)
    ax.set_ylim(1, -max_depth)
    
    # Set y-axis ticks
    y_ticks = []
    y_labels = []
    for y in range(0, int(max_depth) + 1, 2):
        y_ticks.append(-y)
        y_labels.append(f'-{y}' if y > 0 else '0')
    ax.set_yticks(y_ticks)
    ax.set_yticklabels(y_labels)
    ax.set_ylabel('Depth (m)', fontsize=12)
    ax.invert_yaxis()
    
    # Draw soil layers (refer to web logic)
    prev_depth = 0
    for layer in layers:
        bot_depth = float(layer.get('bot_depth', 0))
        layer_type = str(layer.get('type', 'D')).strip().upper()
        
        # U layer: dark brown; D layer: yellow
        if layer_type == 'U' or layer_type == 'UNDRAINED':
            rgba_color = (0.6, 0.4, 0.2, 0.6)
        else:
            rgba_color = (0.941, 0.902, 0.549, 0.5)
        
        # Determine relationship between soil layer and retaining wall depth
        layer_top = prev_depth
        layer_bottom = bot_depth
        
        # If layer crosses wall depth, draw in two parts
        if layer_top < wall_length and layer_bottom > wall_length:
            # Upper part: above wall depth
            # Left (excavation side): draw soil only below excavation depth
            if layer_bottom > excavation_depth:
                left_top = -max(layer_top, excavation_depth)
                left_bottom = -wall_length
                if left_top > left_bottom:
                    rect = patches.Rectangle((0, left_bottom), 4.85, left_top - left_bottom,
                                           linewidth=1, edgecolor='#888', facecolor=rgba_color)
                    ax.add_patch(rect)
            
            # Right (non-excavation side): draw fully
            right_top = -layer_top
            right_bottom = -wall_length
            if right_top > right_bottom:
                rect = patches.Rectangle((5.15, right_bottom), 4.85, right_top - right_bottom,
                                       linewidth=1, edgecolor='#888', facecolor=rgba_color)
                ax.add_patch(rect)
            
            # Lower part: below wall depth, horizontally continuous
            lower_top = -wall_length
            lower_bottom = -layer_bottom
            if lower_top > lower_bottom:
                rect = patches.Rectangle((0, lower_bottom), 10, lower_top - lower_bottom,
                                       linewidth=1, edgecolor='#888', facecolor=rgba_color)
                ax.add_patch(rect)
        elif layer_bottom > wall_length:
            # Layer entirely below wall depth; draw horizontally continuous
            continuous_top = -layer_top
            continuous_bottom = -layer_bottom
            if continuous_top > continuous_bottom:
                rect = patches.Rectangle((0, continuous_bottom), 10, continuous_top - continuous_bottom,
                                       linewidth=1, edgecolor='#888', facecolor=rgba_color)
                ax.add_patch(rect)
        else:
            # Layer entirely above wall depth; draw left and right separately
            # Left (excavation side): draw soil only below excavation depth
            if layer_bottom > excavation_depth:
                left_top = -max(layer_top, excavation_depth)
                left_bottom = -layer_bottom
                if left_top > left_bottom:
                    rect = patches.Rectangle((0, left_bottom), 4.85, left_top - left_bottom,
                                           linewidth=1, edgecolor='#888', facecolor=rgba_color)
                    ax.add_patch(rect)
            
            # Right (non-excavation side): show all layers fully
            right_top = -layer_top
            right_bottom = -layer_bottom
            if right_top > right_bottom:
                rect = patches.Rectangle((5.15, right_bottom), 4.85, right_top - right_bottom,
                                       linewidth=1, edgecolor='#888', facecolor=rgba_color)
                ax.add_patch(rect)
        
        prev_depth = bot_depth
    
    # Draw excavation area (top left, from ground to excavation depth)
    if excavation_depth > 0:
        excavation_rect = patches.Rectangle((0, 0), 4.85, -excavation_depth,
                                           linewidth=2, edgecolor='#333', facecolor='#ddd', alpha=0.3)
        ax.add_patch(excavation_rect)
        
        # Draw excavation surface line
        ax.plot([0, 4.85], [-excavation_depth, -excavation_depth], 'k-', linewidth=3)
    
    # Draw retaining wall
    wall_rect = patches.Rectangle((4.85, -wall_length), 0.3, wall_length,
                                 linewidth=0, facecolor='#666')
    ax.add_patch(wall_rect)
    
    # Draw water level
    if water_levels:
        water_level = water_levels[0] if isinstance(water_levels, list) else water_levels
        if water_level > 0:
            ax.plot([0, 10], [-water_level, -water_level], 'b--', linewidth=2, label='Water Level')
    
    ax.set_title(f'Excavation Profile - {stage_name}', fontsize=14, fontweight='bold')
    ax.grid(False)
    
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
    plt.close(fig)
    img_buffer.seek(0)
    
    return img_buffer


def _rich_sub(txt):
    """Return TextBlock with subscript formatting."""
    return TextBlock(InlineFont(vertAlign='subscript'), txt)


def generate_excavation_excel(project_info, uplift_results, sand_boil_results, layers, metadata=None, stages=None):
    """Generate excavation analysis Excel report; display units per unit_system."""
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
    header_font = Font(name="Times New Roman", bold=True, size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    unit_system = (metadata or {}).get('unit_system', 'metric')
    len_unit = 'ft' if unit_system == 'imperial' else 'm'
    press_unit = 'psf' if unit_system == 'imperial' else 'kPa'
    sheet_index = 0

    # Excavation Profile (metric for plotting) - per-stage excavation
    _gwt_m = float((metadata or {}).get('gwt_gl', 2.0))
    _raw_stages = list(stages or [])
    if not _raw_stages and (uplift_results or sand_boil_results):
        # Infer stages from uplift/sand boil results when request has no stages
        seen = set()
        for iface in (uplift_results or []):
            for st in iface.get('stages', []):
                de = st.get('De')
                if de is not None and de != '-' and (st.get('Stage'), de) not in seen:
                    try:
                        d = float(de)
                        dw = st.get('Dw')
                        wl = float(dw) if dw not in (None, '-') and str(dw).strip() else _gwt_m
                        seen.add((st.get('Stage'), de))
                        _raw_stages.append({'name': st.get('Stage', 'Stage'), 'depth': d, 'water_levels': [wl]})
                    except (TypeError, ValueError):
                        pass
        if not _raw_stages:
            for st in (sand_boil_results or []):
                de = st.get('De')
                if de is not None and de != '-' and (st.get('Stage'), de) not in seen:
                    try:
                        d = float(de)
                        seen.add((st.get('Stage'), de))
                        _raw_stages.append({'name': st.get('Stage', 'Stage'), 'depth': d, 'water_levels': [_gwt_m]})
                    except (TypeError, ValueError):
                        pass
    _raw_layers = layers or []
    _wall_m = float((metadata or {}).get('wall_length', 21.0))
    _if_m = float((metadata or {}).get('interface_depth', 4.0))
    if unit_system == 'imperial':
        _layers_m = [{'bot_depth': float(L.get('bot_depth', L.get('depth', 0))) * FT_TO_M, 'gamma': float(L.get('gamma', L.get('unit_weight', 0))) * PCF_TO_TF_M3, 'type': L.get('type', 'D'), 'code': L.get('code', '')} for L in _raw_layers]
        _stages_m = [{'name': s.get('name', 'Stage'), 'depth': float(s.get('depth', 0)) * FT_TO_M, 'water_levels': [float(w) * FT_TO_M for w in (s.get('water_levels') if isinstance(s.get('water_levels'), list) and s.get('water_levels') else [float(s.get('water_level', _gwt_m))])]} for s in _raw_stages]
        _wall_m *= FT_TO_M
        _gwt_m *= FT_TO_M
        _if_m *= FT_TO_M
    else:
        _layers_m = [{'bot_depth': float(L.get('bot_depth', L.get('depth', 0))), 'gamma': float(L.get('gamma', L.get('unit_weight', 0))), 'type': L.get('type', 'D'), 'code': L.get('code', '')} for L in _raw_layers]
        _stages_m = [{'name': s.get('name', 'Stage'), 'depth': float(s.get('depth', 0)), 'water_levels': [float(w) for w in (s.get('water_levels') if isinstance(s.get('water_levels'), list) and s.get('water_levels') else [float(s.get('water_level', _gwt_m))])]} for s in _raw_stages]

    # Excavation Profile (standalone sheet, first; 2-column layout per DarahSan)
    if _layers_m and _stages_m:
        ws_profile = wb.create_sheet("Excavation Profile", sheet_index)
        sheet_index += 1
        ws_profile.sheet_view.showGridLines = False
        ws_profile.append(["Excavation Profile - Per-stage excavation profiles"])
        ws_profile.append([])
        start_row = 3
        max_layer_depth = max([L.get('bot_depth', 0) for L in _layers_m]) if _layers_m else 0
        max_excavation_depth = max([float(s.get('depth', 0)) for s in _stages_m]) if _stages_m else 0
        global_max_depth = max(max_layer_depth, _wall_m, max_excavation_depth) + 5
        for i, st in enumerate(_stages_m):
            try:
                dep = float(st.get('depth', 0))
                wls = st.get('water_levels', [_gwt_m])
                wl = float(wls[0]) if wls else _gwt_m
                stage_name = st.get('name', 'Stage')
                buf = draw_excavation_profile(_layers_m, _wall_m, dep, [wl], _if_m, stage_name, max_depth=global_max_depth)
                if not buf or len(buf.getvalue()) == 0:
                    continue
                buf.seek(0)
                img = Image(io.BytesIO(buf.getvalue()))
                img.width = 300
                img.height = 375
                row_offset = i // 2
                row_num = start_row + row_offset
                column = 'A' if i % 2 == 0 else 'D'
                ws_profile.row_dimensions[row_num].height = 285
                if row_num > ws_profile.max_row:
                    for col in range(1, 11):
                        ws_profile.cell(row=row_num, column=col).value = None
                img.anchor = f'{column}{row_num}'
                ws_profile.add_image(img)
            except Exception as e:
                import traceback
                print(f"Excavation Profile draw error for stage {i+1}: {e}")
                traceback.print_exc()

    if uplift_results:
        ws_uplift = wb.create_sheet("Uplift", sheet_index)
        sheet_index += 1
        ws_uplift.sheet_view.showGridLines = False
        ws_uplift.append(["Uplift Analysis Results"])
        ws_uplift.append([])
        if metadata:
            ws_uplift.append([CellRichText("Required Safety Factor: FS", _rich_sub("u"), f",r = {metadata.get('fs_u_required', 'N/A')}")])
            ws_uplift.append([])
        ws_uplift.append(["Calculation Formula"])
        ws_uplift.append([])
        formula_font = Font(name="Times New Roman", size=11)
        formula_center = Alignment(horizontal='center', vertical='center')
        uplift_start_row = ws_uplift.max_row + 1
        ws_uplift.cell(row=uplift_start_row, column=1, value=CellRichText("FS", _rich_sub("u"), " =")).font = formula_font
        ws_uplift.cell(row=uplift_start_row, column=2, value=CellRichText("Σ γ", _rich_sub("ti"), " × h", _rich_sub("i"))).font = formula_font
        ws_uplift.cell(row=uplift_start_row, column=2).alignment = formula_center
        ws_uplift.cell(row=uplift_start_row, column=3, value="≥ 1.20").font = formula_font
        ws_uplift.cell(row=uplift_start_row + 1, column=2, value=CellRichText("U", _rich_sub("w"))).font = formula_font
        ws_uplift.cell(row=uplift_start_row + 1, column=2).alignment = formula_center
        ws_uplift.merge_cells(start_row=uplift_start_row, start_column=1, end_row=uplift_start_row + 1, end_column=1)
        ws_uplift.cell(row=uplift_start_row, column=1).alignment = formula_center
        ws_uplift.merge_cells(start_row=uplift_start_row, start_column=3, end_row=uplift_start_row + 1, end_column=3)
        ws_uplift.cell(row=uplift_start_row, column=3).alignment = formula_center
        white_side = Side(style='thin', color='FFFFFF')
        uplift_white_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)
        uplift_frac_top = Border(left=white_side, right=white_side, top=Side(style='thin'), bottom=white_side)
        for r in range(uplift_start_row, uplift_start_row + 2):
            for c in range(1, 4):
                if (r, c) == (uplift_start_row + 1, 2):
                    ws_uplift.cell(row=r, column=c).border = uplift_frac_top
                else:
                    ws_uplift.cell(row=r, column=c).border = uplift_white_border
        ws_uplift.append([])
        headers = ["Stage", f"De (GL {len_unit})", f"Dw (GL {len_unit})", f"Uw ({press_unit})", f"Σγti×hi ({press_unit})", "FSu", "Check Result"]
        ws_uplift.append(headers)
        for cell in ws_uplift[ws_uplift.max_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        for interface_data in uplift_results:
            ws_uplift.append([f"Uplift Interface: {interface_data.get('interface', '')}"])
            ws_uplift.append([])
            for stage_data in interface_data.get('stages', []):
                row = [
                    stage_data.get('Stage', ''),
                    stage_data.get('De', ''),
                    stage_data.get('Dw', ''),
                    stage_data.get('Uw', ''),
                    stage_data.get('Weight', ''),
                    stage_data.get('FSu', ''),
                    stage_data.get('Result', '') or stage_data.get('Note', '') or stage_data.get('Status', '')
                ]
                ws_uplift.append(row)
                for cell in ws_uplift[ws_uplift.max_row]:
                    cell.border = border
            ws_uplift.append([])

    if sand_boil_results:
        ws_sand = wb.create_sheet("Sand Boil", sheet_index)
        sheet_index += 1
        ws_sand.sheet_view.showGridLines = False
        ws_sand.append(["Sand Boil Analysis Results"])
        ws_sand.append([])
        if metadata:
            ws_sand.append([CellRichText("Required Safety Factor: FS", _rich_sub("p1"), f",r = {metadata.get('fs_p1_required', 'N/A')}"), CellRichText("FS", _rich_sub("p2"), f",r = {metadata.get('fs_p2_required', 'N/A')}")])
            ws_sand.append([])
        ws_sand.append(["Calculation Formulas"])
        ws_sand.append([])
        formula_font = Font(name="Times New Roman", size=11)
        formula_center = Alignment(horizontal='center', vertical='center')
        start_row = ws_sand.max_row + 1
        ws_sand.cell(row=start_row, column=1, value=CellRichText("FS", _rich_sub("p1"), " =")).font = formula_font
        ws_sand.cell(row=start_row, column=2, value=CellRichText("2γ", _rich_sub("sub"), " × D")).font = formula_font
        ws_sand.cell(row=start_row, column=2).alignment = formula_center
        ws_sand.cell(row=start_row, column=3, value="≥ 1.50").font = formula_font
        ws_sand.cell(row=start_row + 1, column=2, value=CellRichText("γ", _rich_sub("w"), " ΔH", _rich_sub("w"))).font = formula_font
        ws_sand.cell(row=start_row + 1, column=2).alignment = formula_center
        ws_sand.merge_cells(start_row=start_row, start_column=1, end_row=start_row + 1, end_column=1)
        ws_sand.cell(row=start_row, column=1).alignment = formula_center
        ws_sand.merge_cells(start_row=start_row, start_column=3, end_row=start_row + 1, end_column=3)
        ws_sand.cell(row=start_row, column=3).alignment = formula_center
        ws_sand.cell(row=start_row + 2, column=1, value=CellRichText("FS", _rich_sub("p2"), " =")).font = formula_font
        ws_sand.cell(row=start_row + 2, column=2, value=CellRichText("γ", _rich_sub("sub"), " × (ΔH", _rich_sub("w"), " + 2D)")).font = formula_font
        ws_sand.cell(row=start_row + 2, column=2).alignment = formula_center
        ws_sand.cell(row=start_row + 2, column=3, value="≥ 2.00").font = formula_font
        ws_sand.cell(row=start_row + 3, column=2, value=CellRichText("γ", _rich_sub("w"), " ΔH", _rich_sub("w"))).font = formula_font
        ws_sand.cell(row=start_row + 3, column=2).alignment = formula_center
        ws_sand.merge_cells(start_row=start_row + 2, start_column=1, end_row=start_row + 3, end_column=1)
        ws_sand.cell(row=start_row + 2, column=1).alignment = formula_center
        ws_sand.merge_cells(start_row=start_row + 2, start_column=3, end_row=start_row + 3, end_column=3)
        ws_sand.cell(row=start_row + 2, column=3).alignment = formula_center
        # Keep fraction line; set other borders to white
        white_side = Side(style='thin', color='FFFFFF')
        formula_white_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)
        formula_frac_top = Border(left=white_side, right=white_side, top=Side(style='thin'), bottom=white_side)  # Fraction line between numerator and denominator
        for r in range(start_row, start_row + 4):
            for c in range(1, 4):
                if (r, c) in ((start_row + 1, 2), (start_row + 3, 2)):
                    ws_sand.cell(row=r, column=c).border = formula_frac_top  # Denominator cell: keep top line as fraction bar
                else:
                    ws_sand.cell(row=r, column=c).border = formula_white_border
        ws_sand.append([])
        # D = excavation surface to retaining wall bottom (embedment depth); computed per stage
        ws_sand.append([f"D = Excavation surface to retaining wall bottom ({len_unit})"])
        ws_sand.cell(row=ws_sand.max_row, column=1).font = Font(name="Times New Roman", italic=True, size=10)
        ws_sand.append([])
        headers = ["Excavation Stage", f"De (GL {len_unit})", "", f"D ({len_unit})",
                   "", "", "", "", "Check Result"]
        ws_sand.append(headers)
        header_row = ws_sand.max_row
        ws_sand.cell(row=header_row, column=3, value=CellRichText("ΔH", _rich_sub("w"), f" ({len_unit})"))
        ws_sand.cell(row=header_row, column=5, value=CellRichText("γ", _rich_sub("sub"), f"D ({press_unit})"))
        ws_sand.cell(row=header_row, column=6, value=CellRichText("γ", _rich_sub("sub"), "ΔH", _rich_sub("w"), f" ({press_unit})"))
        ws_sand.cell(row=header_row, column=7, value=CellRichText("FS", _rich_sub("p1")))
        ws_sand.cell(row=header_row, column=8, value=CellRichText("FS", _rich_sub("p2")))
        for cell in ws_sand[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        def _fmt(val):
            """Format value for Excel: round numbers to 2 decimals, keep '-' as-is."""
            if val == '-' or val is None:
                return '-'
            try:
                f = float(val)
                return round(f, 2) if isinstance(val, (int, float)) else val
            except (TypeError, ValueError):
                return val

        for result in sand_boil_results:
            check_result = result.get('Result', '') or result.get('Note', '') or result.get('Status', '')
            row = [
                result.get('Stage', ''),
                _fmt(result.get('De', '')),
                _fmt(result.get('dHw', '')),
                _fmt(result.get('D', '')),  # D (embedment) computed per stage
                _fmt(result.get('gamma_sub_D', '')),
                _fmt(result.get('gamma_sub_delta_hw', '')),
                _fmt(result.get('FSp1', '')),
                _fmt(result.get('FSp2', '')),
                check_result
            ]
            ws_sand.append(row)
            for cell in ws_sand[ws_sand.max_row]:
                cell.border = border
                if cell.column == 9:
                    v = str(cell.value or '').upper()
                    if 'NG' in v or 'FAIL' in v:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif 'OK' in v or 'NO CHECK' in v.upper():
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws_symbols = wb.create_sheet("Symbol Description", sheet_index)
    ws_symbols.sheet_view.showGridLines = False
    symbol_font = Font(name="Times New Roman", size=11)
    symbol_header_font = Font(name="Times New Roman", bold=True, size=11)
    ws_symbols.column_dimensions["B"].width = 48
    ws_symbols.append(["Symbol Description"])
    ws_symbols.append([])
    ws_symbols.append(["Symbol", "Description"])
    ws_symbols.append(["De", f"Excavation Depth ({len_unit})"])
    ws_symbols.append(["Dw", f"Water Level ({len_unit})"])
    ws_symbols.append(["Uw", f"Water Pressure ({press_unit})"])
    ws_symbols.append(["Σγti×hi", f"Total Weight ({press_unit})"])
    ws_symbols.append([CellRichText("FS", _rich_sub("u")), "Factor of Safety for Uplift"])
    ws_symbols.append([CellRichText("FS", _rich_sub("p1")), "Factor of Safety for Sand Boil (Formula 1)"])
    ws_symbols.append([CellRichText("FS", _rich_sub("p2")), "Factor of Safety for Sand Boil (Formula 2)"])
    ws_symbols.append([CellRichText("γ", _rich_sub("sub")), f"Submerged Unit Weight ({'pcf' if unit_system == 'imperial' else 'kN/m³'})"])
    ws_symbols.append(["D", f"Embedment depth: excavation surface to retaining wall bottom ({len_unit})"])
    ws_symbols.append([CellRichText("ΔH", _rich_sub("w")), f"Water Head Difference ({len_unit})"])
    # Apply black table borders and Times New Roman font for Symbol Description body.
    symbol_table_start = 3
    symbol_table_end = ws_symbols.max_row
    for r in range(symbol_table_start, symbol_table_end + 1):
        for c in range(1, 3):
            cell = ws_symbols.cell(row=r, column=c)
            cell.border = border
            cell.font = symbol_header_font if r == symbol_table_start else symbol_font
            if r == symbol_table_start:
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    for r in range(1, symbol_table_start):
        for c in range(1, 4):
            ws_symbols.cell(row=r, column=c).font = symbol_header_font if r == 1 else symbol_font

    for ws in wb.worksheets:
        for col in range(1, 12):
            col_letter = get_column_letter(col)
            if col_letter not in ['A', 'D']:
                ws.column_dimensions[col_letter].width = 15
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer
